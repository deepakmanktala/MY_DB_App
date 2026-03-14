# =============================================================================
# ML PROGRAMS WITH DETAILED COMMENTS + CHARTS EMBEDDED IN EXCEL
# Covers: Linear Regression, Logistic Regression, Underfitting, Overfitting
# =============================================================================

# --- IMPORTS -----------------------------------------------------------------
# 'import' keyword tells Python to load an external library into memory
# so we can use its functions and classes in our program

import numpy as np
# numpy  = "Numerical Python" — the core library for numerical computing
# np     = alias (shortcut name) so we type np.array() instead of numpy.array()
# provides: arrays, math ops, random number generation, linear algebra

import matplotlib
# matplotlib = the standard Python plotting/charting library
# handles: line charts, scatter plots, bar charts, histograms, etc.

matplotlib.use('Agg')
# .use('Agg') = set the rendering backend BEFORE importing pyplot
# 'Agg' = Anti-Grain Geometry backend — renders to image files (PNG/JPG)
# required because we are running headless (no display screen attached)
# without this, matplotlib tries to open a GUI window and crashes on servers

import matplotlib.pyplot as plt
# pyplot = the submodule of matplotlib that provides MATLAB-style plotting API
# plt    = standard alias — lets us call plt.plot(), plt.show() etc.

import matplotlib.patches as mpatches
# patches = submodule for drawing shapes (rectangles, arrows, circles)
# mpatches = alias — used here to create custom legend entries with color boxes

from io import BytesIO
# io       = Python's built-in module for working with data streams
# BytesIO  = an in-memory binary stream — acts like a file but lives in RAM
# we use it to save chart images to memory instead of disk, then embed in Excel

from sklearn.linear_model import LinearRegression, LogisticRegression
# sklearn          = scikit-learn: the standard Python machine learning library
# linear_model     = submodule containing regression and classification models
# LinearRegression = fits a straight line (or hyperplane) to continuous data
# LogisticRegression = despite the name, used for CLASSIFICATION (not regression)
#                      outputs a probability via the sigmoid function

from sklearn.preprocessing import PolynomialFeatures
# preprocessing     = submodule for transforming/preparing data before modeling
# PolynomialFeatures = transforms input features [x] into [1, x, x², x³, ...]
#                      lets a linear model fit non-linear (curved) patterns

from sklearn.tree import DecisionTreeClassifier
# tree                    = submodule for tree-based models
# DecisionTreeClassifier  = builds a flowchart-like tree of if/else rules
#                           each split tries to separate classes as cleanly as possible

from sklearn.model_selection import train_test_split
# model_selection   = submodule for model evaluation utilities
# train_test_split  = shuffles dataset and splits into two parts:
#                     training set (model learns from this)
#                     test set     (model is evaluated on this — never seen during training)

from sklearn.metrics import (
    mean_squared_error,   # MSE: average of (actual - predicted)² — for regression
    r2_score,             # R²:  proportion of variance explained (1.0 = perfect)
    accuracy_score,       # Accuracy: fraction of correct predictions — for classification
    confusion_matrix      # Confusion matrix: table of TP, TN, FP, FN counts
)

from openpyxl import Workbook
# openpyxl  = library for creating and editing Excel (.xlsx) files in Python
# Workbook  = represents an entire Excel file — contains one or more worksheets

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
# styles    = submodule for cell formatting in Excel
# Font      = controls text appearance: bold, italic, size, color, font family
# PatternFill = fills a cell background with a solid color or pattern
# Alignment = controls text position: left/center/right, top/middle/bottom, wrap
# Border    = defines border lines around cells
# Side      = defines one side of a border: style (thin/thick) and color

from openpyxl.utils import get_column_letter
# utils           = utility functions for openpyxl
# get_column_letter = converts column number to Excel letter
#                     e.g. get_column_letter(1) → 'A', get_column_letter(27) → 'AA'

from openpyxl.drawing.image import Image as XLImage
# drawing.image  = submodule for inserting images into Excel worksheets
# Image          = class that wraps an image file so it can be placed on a sheet
# aliased XLImage to avoid conflict with Python's built-in 'Image' name

import warnings
# warnings = Python's built-in module for controlling warning messages
warnings.filterwarnings('ignore')
# filterwarnings('ignore') = suppress all warning messages
# prevents sklearn convergence warnings from cluttering the output

# =============================================================================
# GLOBAL RANDOM SEED — REPRODUCIBILITY
# =============================================================================

np.random.seed(42)
# np.random.seed() = initialises the random number generator with a fixed value
# seed=42          = arbitrary number — ensures SAME random numbers every run
# without this, numpy generates different random numbers each run
# → results would differ each time the program is executed
# 42 is a popular convention (from "Hitchhiker's Guide to the Galaxy")

# =============================================================================
# STYLING HELPER FUNCTIONS
# These functions are defined once and reused across all sheets
# 'def' keyword = define a new function; 'return' = send value back to caller
# =============================================================================

def hex_fill(hex_color):
    # PatternFill creates a solid background fill for an Excel cell
    # fgColor = foreground color = the fill color (confusingly named in openpyxl)
    # 'solid' = fill type: fills entire cell with one color (no pattern/gradient)
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    # Side defines one border edge: style='thin' = 1px line, color in hex RRGGBB
    s = Side(style='thin', color='CCCCCC')
    # Border applies the same thin side to all four edges of the cell
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_font(color="FFFFFF", size=10, bold=True):
    # Font controls text styling inside a cell
    # name='Arial'  = font family
    # bold=True     = makes text bold (thicker strokes)
    # color=color   = hex color string for the text (default white for dark headers)
    # size=size     = font size in points
    return Font(name='Arial', bold=bold, color=color, size=size)

def body_font(color="1A1A2E", size=9, bold=False):
    # same as hdr_font but defaulting to dark text on light backgrounds
    return Font(name='Arial', bold=bold, color=color, size=size)

def center():
    # Alignment controls where text sits inside the cell
    # horizontal='center' = horizontally centred
    # vertical='center'   = vertically centred (important when row is tall)
    return Alignment(horizontal='center', vertical='center')

def left_indent():
    # wrap_text=True = text wraps to next line if it exceeds column width
    # indent=1       = adds a small left padding so text doesn't touch cell edge
    return Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)

def style_header(ws, row, col, text, bg, fg="FFFFFF", colspan=1, height=22):
    # ws        = worksheet object to write into
    # row, col  = 1-based row and column index of the cell
    # text      = string to display
    # bg        = background hex color
    # fg        = foreground (text) hex color
    # colspan   = how many columns to merge (1 = no merge)
    # height    = row height in points
    c = ws.cell(row=row, column=col, value=text)
    # .cell(row, column, value) = write 'text' into the specified cell
    c.font      = hdr_font(color=fg)          # apply header font styling
    c.fill      = hex_fill(bg)                # apply background color
    c.alignment = left_indent()               # left-align with indent
    c.border    = thin_border()               # add thin border around cell
    ws.row_dimensions[row].height = height   # set row height in points
    if colspan > 1:
        # merge_cells combines multiple cells into one visual cell
        # the combined cell spans from (row,col) to (row, col+colspan-1)
        end_col = get_column_letter(col + colspan - 1)  # convert number → letter
        start_col = get_column_letter(col)
        ws.merge_cells(f'{start_col}{row}:{end_col}{row}')
    return c   # return the cell so the caller can further modify it if needed

def write_kv(ws, row, col, key, val, key_bg="E8EAF6", val_bg="FFFFFF"):
    # Writes a key-value pair side by side in two cells
    # Used for displaying model metrics like "R² Train | 0.9423"
    k = ws.cell(row=row, column=col, value=key)
    k.font      = body_font(bold=True)   # bold key text
    k.fill      = hex_fill(key_bg)       # light purple-grey background for keys
    k.alignment = left_indent()
    k.border    = thin_border()
    ws.row_dimensions[row].height = 18   # 18pt row height for readability

    v = ws.cell(row=row, column=col + 1, value=val)
    v.font      = body_font()            # normal weight for value text
    v.fill      = hex_fill(val_bg)       # white background for values
    v.alignment = center()
    v.border    = thin_border()

def set_col_widths(ws, widths):
    # Sets the width of each column in the worksheet
    # widths = list of widths in Excel column-width units (approx. characters wide)
    # enumerate(widths, 1) = iterate over widths list, starting column index at 1
    for col_idx, width in enumerate(widths, 1):
        letter = get_column_letter(col_idx)  # convert 1→'A', 2→'B', etc.
        ws.column_dimensions[letter].width = width  # set column width

def insert_chart_image(ws, img_bytes, anchor_cell, width_px=480, height_px=320):
    # Embeds a matplotlib chart PNG image directly into an Excel worksheet
    # img_bytes   = BytesIO object containing the PNG image data in memory
    # anchor_cell = e.g. 'G2' — top-left corner where the image will be placed
    # width_px    = display width of image in Excel (in pixels/EMUs)
    # height_px   = display height of image in Excel
    img_bytes.seek(0)
    # .seek(0) = rewind the BytesIO stream cursor to the beginning
    # BytesIO works like a file — after writing, the cursor is at the end
    # we must rewind to start so openpyxl can READ from the beginning
    img = XLImage(img_bytes)           # wrap BytesIO in openpyxl Image object
    img.width  = width_px              # set display width in Excel
    img.height = height_px             # set display height in Excel
    ws.add_image(img, anchor_cell)     # place image at the specified cell anchor

# =============================================================================
# CHART GENERATION FUNCTIONS
# Each function creates one matplotlib figure, saves to BytesIO, returns it
# =============================================================================

def chart_linear_regression(X_train, y_train, X_test, y_test,
                            y_pred_train, y_pred_test, coef, intercept):
    """
    Creates a scatter + regression line chart for Linear Regression.
    Shows training points (blue), test points (red), and the fitted line (green).
    """
    fig, axes = plt.subplots(1, 2, figsize=(12, 5))
    # plt.subplots(rows, cols, figsize) = create a figure with a grid of subplots
    # 1 row, 2 cols = side-by-side charts
    # figsize=(12,5) = figure is 12 inches wide × 5 inches tall at 100 DPI

    fig.patch.set_facecolor('#0D1117')
    # fig.patch = the figure background rectangle
    # set_facecolor = sets background color using a hex string
    # '#0D1117' = very dark navy — matches our Excel dark theme

    for ax in axes:
        # ax = one subplot (Axes object); iterate both to apply same background
        ax.set_facecolor('#161B22')   # dark background for the plot area
        ax.tick_params(colors='#C9D1D9', labelsize=8)
        # tick_params = style the axis tick marks and their labels
        # colors = color of tick marks and tick label text
        # labelsize = font size of tick labels in points
        ax.spines['bottom'].set_color('#30363D')  # x-axis line color
        ax.spines['left'].set_color('#30363D')    # y-axis line color
        ax.spines['top'].set_visible(False)       # hide top border line
        ax.spines['right'].set_visible(False)     # hide right border line
        # spines = the four border lines around the plot area

    # --- Left chart: training data + fitted line ---
    ax1 = axes[0]   # axes[0] = first (left) subplot
    ax1.scatter(X_train, y_train, color='#58A6FF', s=60, alpha=0.8,
                zorder=3, label='Train data')
    # scatter() = draws individual data points as dots
    # X_train   = x-axis values (house sizes)
    # y_train   = y-axis values (house prices)
    # color     = dot color (blue)
    # s=60      = dot size in points²
    # alpha=0.8 = 80% opacity (slight transparency)
    # zorder=3  = drawing layer order (higher = drawn on top)
    # label     = text for legend entry

    x_line = np.linspace(X_train.min(), X_train.max(), 100)
    # np.linspace(start, stop, num) = creates 100 evenly spaced x values
    # used to draw a smooth continuous line (not just connect data dots)
    y_line = intercept + coef * x_line
    # calculate predicted y for each x point using the learned equation:
    # y = β₀ + β₁·x   (intercept + coefficient × x)

    ax1.plot(x_line, y_line, color='#3FB950', linewidth=2, zorder=2, label='Fitted line')
    # plot()     = draws a line connecting (x,y) pairs
    # linewidth=2 = line thickness in points
    ax1.set_title('Training Data + Regression Line', color='#E6EDF3', fontsize=10, pad=10)
    ax1.set_xlabel('House Size (sqft)', color='#8B949E', fontsize=9)
    ax1.set_ylabel('Price ($k)', color='#8B949E', fontsize=9)
    ax1.legend(facecolor='#21262D', edgecolor='#30363D', labelcolor='#C9D1D9', fontsize=8)
    # legend() = shows the color-coded key
    # facecolor = legend box background
    # edgecolor = legend box border
    # labelcolor = text color in the legend

    # --- Right chart: test data vs predictions ---
    ax2 = axes[1]   # axes[1] = second (right) subplot
    ax2.scatter(X_test, y_test, color='#F85149', s=60, alpha=0.8,
                zorder=3, label='Actual test values')
    ax2.scatter(X_test, y_pred_test, color='#58A6FF', s=60, alpha=0.8,
                marker='^', zorder=4, label='Predicted values')
    # marker='^' = triangle marker shape (instead of default circle)

    # draw vertical error lines connecting actual to predicted for each test point
    for xv, ya, yp in zip(X_test, y_test, y_pred_test):
        # zip() = pairs up corresponding elements from three lists simultaneously
        ax2.plot([xv, xv], [ya, yp], color='#FFD700', alpha=0.4, linewidth=1)
        # [xv, xv] = same x for both endpoints → vertical line
        # [ya, yp] = y goes from actual to predicted → shows the error/residual

    ax2.set_title('Test Set: Actual vs Predicted', color='#E6EDF3', fontsize=10, pad=10)
    ax2.set_xlabel('House Size (sqft)', color='#8B949E', fontsize=9)
    ax2.set_ylabel('Price ($k)', color='#8B949E', fontsize=9)
    ax2.legend(facecolor='#21262D', edgecolor='#30363D', labelcolor='#C9D1D9', fontsize=8)

    plt.tight_layout(pad=2.0)
    # tight_layout() = automatically adjusts subplot spacing to prevent overlap
    # pad=2.0 = padding in inches between subplots and figure edges

    buf = BytesIO()
    # BytesIO() = creates an empty in-memory binary buffer (acts like an open file)
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight',
                facecolor=fig.get_facecolor())
    # savefig()         = renders the figure and writes it to the buffer
    # format='png'      = save as PNG image format
    # dpi=130           = dots per inch — higher = sharper image
    # bbox_inches='tight' = crop whitespace around the figure
    # facecolor         = preserve the dark background color in the saved image
    plt.close(fig)
    # close(fig) = release memory used by the figure — important in loops
    return buf   # return BytesIO buffer containing the PNG bytes


def chart_logistic_regression(X_test, y_test, y_pred, y_prob):
    """
    Creates two charts for Logistic Regression:
    Left:  Scatter of study hours vs probability of passing (sigmoid curve shape)
    Right: Bar chart showing predicted probability for each test student
    """
    fig, axes = plt.subplots(1, 2, figsize=(12, 5))
    fig.patch.set_facecolor('#0D1117')

    for ax in axes:
        ax.set_facecolor('#161B22')
        ax.tick_params(colors='#C9D1D9', labelsize=8)
        ax.spines['bottom'].set_color('#30363D')
        ax.spines['left'].set_color('#30363D')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

    # --- Left chart: study hours vs pass probability ---
    ax1 = axes[0]
    # separate test students into pass and fail groups for different colors
    pass_mask = y_test == 1    # boolean array: True where student passed
    fail_mask = y_test == 0    # boolean array: True where student failed
    # boolean indexing: X_test[pass_mask, 0] = study hours of passing students only

    ax1.scatter(X_test[pass_mask, 0], y_prob[pass_mask],
                color='#3FB950', s=70, alpha=0.85, label='Actual: Pass', zorder=3)
    # X_test[:, 0] = first column (study hours); [:, 1] would be second (sleep hours)
    ax1.scatter(X_test[fail_mask, 0], y_prob[fail_mask],
                color='#F85149', s=70, alpha=0.85, label='Actual: Fail', zorder=3)

    ax1.axhline(y=0.5, color='#FFD700', linestyle='--', linewidth=1.5, alpha=0.7,
                label='Decision threshold (0.5)')
    # axhline() = draws a horizontal line across the entire plot
    # y=0.5 = the decision boundary: predict Pass if prob ≥ 0.5
    # linestyle='--' = dashed line
    ax1.fill_between([X_test[:,0].min(), X_test[:,0].max()], 0.5, 1.0,
                     alpha=0.05, color='#3FB950')
    # fill_between() = shades the area between two curves
    # shades the region above 0.5 (the "predict pass" zone) in green

    ax1.set_title('Study Hours vs P(Pass)', color='#E6EDF3', fontsize=10, pad=10)
    ax1.set_xlabel('Study Hours', color='#8B949E', fontsize=9)
    ax1.set_ylabel('Predicted Probability of Passing', color='#8B949E', fontsize=9)
    ax1.set_ylim(0, 1)   # y-axis range fixed 0 to 1 (probabilities always in this range)
    ax1.legend(facecolor='#21262D', edgecolor='#30363D', labelcolor='#C9D1D9', fontsize=8)

    # --- Right chart: per-student probability bars ---
    ax2 = axes[1]
    n = len(y_prob)    # number of test students
    x_pos = np.arange(n)
    # np.arange(n) = creates array [0, 1, 2, ..., n-1]
    # used as x-positions for each bar in the bar chart

    bar_colors = ['#3FB950' if p >= 0.5 else '#F85149' for p in y_prob]
    # list comprehension = concise way to build a list with a condition
    # for each probability p: green if ≥ 0.5 (predict pass), red if < 0.5 (predict fail)

    bars = ax2.bar(x_pos, y_prob, color=bar_colors, alpha=0.85, width=0.7)
    # bar() = creates a bar chart
    # x_pos   = x-axis positions of bars
    # y_prob  = heights of bars (probabilities)
    # color   = list of colors, one per bar
    # width=0.7 = bar width (0 to 1 scale, 1 = bars touching)

    ax2.axhline(y=0.5, color='#FFD700', linestyle='--', linewidth=1.5, alpha=0.8)
    # threshold line at 0.5 — bars above → predict pass, below → predict fail

    # mark bars where prediction was WRONG with an 'X' marker
    for i, (prob, actual, pred) in enumerate(zip(y_prob, y_test, y_pred)):
        if actual != pred:   # prediction was wrong
            ax2.text(i, prob + 0.03, '✗', color='#FFD700', ha='center',
                     fontsize=9, fontweight='bold')
            # ax2.text(x, y, string) = places text at position (x,y) on the chart
            # ha='center' = horizontal alignment centred on x

    ax2.set_title('Predicted Probability per Student', color='#E6EDF3', fontsize=10, pad=10)
    ax2.set_xlabel('Student Index', color='#8B949E', fontsize=9)
    ax2.set_ylabel('P(Pass)', color='#8B949E', fontsize=9)
    ax2.set_ylim(0, 1.15)   # extra space above 1.0 for the ✗ markers

    pass_patch = mpatches.Patch(color='#3FB950', label='Predict: Pass (≥0.5)')
    fail_patch = mpatches.Patch(color='#F85149', label='Predict: Fail (<0.5)')
    # mpatches.Patch = creates a colored rectangle for use as a custom legend entry
    ax2.legend(handles=[pass_patch, fail_patch],
               facecolor='#21262D', edgecolor='#30363D', labelcolor='#C9D1D9', fontsize=8)

    plt.tight_layout(pad=2.0)
    buf = BytesIO()
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight',
                facecolor=fig.get_facecolor())
    plt.close(fig)
    return buf


def chart_underfitting(X_train, y_train, X_test, y_test,
                       y_under_tr, y_under_te, y_good_tr, y_good_te):
    """
    Creates two charts illustrating underfitting:
    Left:  Training data with underfit line vs good-fit curve
    Right: Residual comparison (actual minus predicted) for both models
    """
    fig, axes = plt.subplots(1, 2, figsize=(12, 5))
    fig.patch.set_facecolor('#0D1117')

    for ax in axes:
        ax.set_facecolor('#161B22')
        ax.tick_params(colors='#C9D1D9', labelsize=8)
        ax.spines['bottom'].set_color('#30363D')
        ax.spines['left'].set_color('#30363D')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

    # --- Left chart: data points + two model curves ---
    ax1 = axes[0]
    sort_idx = np.argsort(X_train)
    # np.argsort() = returns the indices that would sort the array
    # e.g. X_train = [3,1,2] → argsort returns [1,2,0] (index of smallest to largest)
    # needed to draw smooth connected lines (must go left to right)

    ax1.scatter(X_train, y_train, color='#79C0FF', s=40, alpha=0.7,
                zorder=2, label='Training data')
    ax1.scatter(X_test, y_test, color='#FFB86C', s=40, alpha=0.7,
                zorder=2, marker='s', label='Test data')
    # marker='s' = square marker shape

    ax1.plot(X_train[sort_idx], y_under_tr[sort_idx],
             color='#F85149', linewidth=2.5, zorder=3, label='Underfit (degree 1)')
    # plotting sorted x vs sorted predictions draws a smooth left-to-right line

    ax1.plot(X_train[sort_idx], y_good_tr[sort_idx],
             color='#3FB950', linewidth=2.5, zorder=4, label='Good fit (degree 3)')

    ax1.set_title('Underfitting: Linear vs Cubic Fit', color='#E6EDF3', fontsize=10, pad=10)
    ax1.set_xlabel('x', color='#8B949E', fontsize=9)
    ax1.set_ylabel('y', color='#8B949E', fontsize=9)
    ax1.legend(facecolor='#21262D', edgecolor='#30363D', labelcolor='#C9D1D9', fontsize=8)

    # --- Right chart: residuals (errors) for each data point ---
    ax2 = axes[1]
    residuals_under = y_test - y_under_te
    # residual = actual value minus predicted value
    # positive residual = model predicted too low
    # negative residual = model predicted too high
    residuals_good  = y_test - y_good_te

    x_idx = np.arange(len(y_test))   # index positions for x-axis
    ax2.bar(x_idx - 0.2, residuals_under, width=0.35, color='#F85149',
            alpha=0.8, label='Underfit residuals')
    # x_idx - 0.2 = shift underfit bars slightly LEFT so both sets are visible side by side
    ax2.bar(x_idx + 0.2, residuals_good,  width=0.35, color='#3FB950',
            alpha=0.8, label='Good-fit residuals')
    # x_idx + 0.2 = shift good-fit bars slightly RIGHT

    ax2.axhline(y=0, color='#8B949E', linewidth=1, alpha=0.6)
    # horizontal line at y=0 = perfect prediction line (zero error)
    # bars above line = under-predicted; bars below line = over-predicted

    ax2.set_title('Residuals: Underfit vs Good Fit', color='#E6EDF3', fontsize=10, pad=10)
    ax2.set_xlabel('Test Sample Index', color='#8B949E', fontsize=9)
    ax2.set_ylabel('Residual (Actual − Predicted)', color='#8B949E', fontsize=9)
    ax2.legend(facecolor='#21262D', edgecolor='#30363D', labelcolor='#C9D1D9', fontsize=8)

    plt.tight_layout(pad=2.0)
    buf = BytesIO()
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight',
                facecolor=fig.get_facecolor())
    plt.close(fig)
    return buf


def chart_overfitting(acc_over_tr, acc_over_te, acc_good_tr, acc_good_te,
                      X_train, y_train, X_test, y_test,
                      over_pred_tr, good_pred_tr):
    """
    Creates two charts illustrating overfitting:
    Left:  Bar chart comparing train vs test accuracy for both models
    Right: Training error pattern — overfit model vs good model
    """
    fig, axes = plt.subplots(1, 2, figsize=(12, 5))
    fig.patch.set_facecolor('#0D1117')

    for ax in axes:
        ax.set_facecolor('#161B22')
        ax.tick_params(colors='#C9D1D9', labelsize=8)
        ax.spines['bottom'].set_color('#30363D')
        ax.spines['left'].set_color('#30363D')
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

    # --- Left chart: accuracy comparison bar chart ---
    ax1 = axes[0]
    categories = ['Overfit Tree\n(depth=∞)', 'Good Tree\n(depth=4)']
    # '\n' in a string = newline character — forces a line break in the label
    train_accs = [acc_over_tr * 100, acc_good_tr * 100]   # convert 0-1 to 0-100%
    test_accs  = [acc_over_te * 100, acc_good_te * 100]

    x = np.arange(len(categories))    # [0, 1] — x positions for the two groups
    width = 0.3                        # width of each individual bar

    bars1 = ax1.bar(x - width/2, train_accs, width, color='#58A6FF',
                    alpha=0.9, label='Train Accuracy')
    # x - width/2 = shift train bars left of centre position
    bars2 = ax1.bar(x + width/2, test_accs,  width, color='#F85149',
                    alpha=0.9, label='Test Accuracy')
    # x + width/2 = shift test bars right of centre position

    # annotate each bar with its exact percentage value
    for bar in list(bars1) + list(bars2):
        # list(bars1) + list(bars2) = combine both BarContainer iterables into one list
        height = bar.get_height()   # get the height (accuracy value) of this bar
        ax1.annotate(f'{height:.1f}%',
                     xy=(bar.get_x() + bar.get_width() / 2, height),
                     xytext=(0, 4),          # offset text 4 points above bar top
                     textcoords='offset points',
                     ha='center', va='bottom',
                     color='#C9D1D9', fontsize=8, fontweight='bold')
        # annotate() = places text at a specific chart position with optional arrow
        # xy = the point being annotated (top-centre of each bar)
        # xytext = where the text is placed (offset from xy in points)

    ax1.set_ylim(0, 115)   # y-axis from 0 to 115% to leave room for text annotations
    ax1.set_xticks(x)      # set x-axis tick positions
    ax1.set_xticklabels(categories, color='#C9D1D9', fontsize=9)
    ax1.set_ylabel('Accuracy (%)', color='#8B949E', fontsize=9)
    ax1.set_title('Overfitting: Train vs Test Accuracy', color='#E6EDF3', fontsize=10, pad=10)
    ax1.legend(facecolor='#21262D', edgecolor='#30363D', labelcolor='#C9D1D9', fontsize=8)

    # draw a horizontal line highlighting the train-test gap for overfit model
    ax1.annotate('', xy=(x[0]+width/2, acc_over_te*100),
                 xytext=(x[0]+width/2, acc_over_tr*100),
                 arrowprops=dict(arrowstyle='<->', color='#FFD700', lw=1.5))
    # arrowprops = draws a double-headed arrow between two points
    # arrowstyle='<->' = arrow with heads at both ends
    # highlights the GAP between train and test accuracy

    # --- Right chart: correct vs wrong predictions per training sample ---
    ax2 = axes[1]
    correct_over = (over_pred_tr == y_train).astype(int)
    # == operator: element-wise comparison → array of True/False
    # .astype(int) converts True→1, False→0
    # 1 = correct prediction, 0 = wrong prediction

    correct_good = (good_pred_tr == y_train).astype(int)

    n = len(y_train)
    ax2.scatter(range(n), correct_over + 0.05,
                c=['#3FB950' if v == 1 else '#F85149' for v in correct_over],
                s=30, alpha=0.8, marker='o', label='Overfit tree')
    # c= (not color=) accepts a list of colors, one per point
    # list comprehension: green if correct (1), red if wrong (0)
    # + 0.05 = slight vertical offset so points don't overlap

    ax2.scatter(range(n), correct_good - 0.05,
                c=['#3FB950' if v == 1 else '#F85149' for v in correct_good],
                s=30, alpha=0.8, marker='^', label='Good tree')

    ax2.set_yticks([0, 1])
    ax2.set_yticklabels(['Wrong', 'Correct'], color='#C9D1D9')
    # set_yticks + set_yticklabels = replace 0/1 numeric labels with 'Wrong'/'Correct'
    ax2.set_title('Training Predictions: Overfit vs Good', color='#E6EDF3', fontsize=10, pad=10)
    ax2.set_xlabel('Training Sample Index', color='#8B949E', fontsize=9)
    ax2.legend(facecolor='#21262D', edgecolor='#30363D', labelcolor='#C9D1D9', fontsize=8)
    ax2.set_ylim(-0.4, 1.4)   # extend y-axis slightly to show offset points cleanly

    plt.tight_layout(pad=2.0)
    buf = BytesIO()
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight',
                facecolor=fig.get_facecolor())
    plt.close(fig)
    return buf


def chart_bias_variance():
    """
    Creates the classic Bias-Variance Tradeoff curve:
    Shows how bias decreases and variance increases as model complexity grows,
    and how total error has a minimum (the 'sweet spot').
    """
    fig, ax = plt.subplots(figsize=(9, 5))
    # single subplot this time (no grid)
    fig.patch.set_facecolor('#0D1117')
    ax.set_facecolor('#161B22')
    ax.tick_params(colors='#C9D1D9', labelsize=9)
    for spine in ax.spines.values():
        spine.set_color('#30363D')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)

    x = np.linspace(0.5, 10, 200)
    # linspace(0.5, 10, 200) = 200 x values from 0.5 to 10
    # represents model complexity from simple (0.5) to very complex (10)

    bias_sq  = 4.0 / x**1.5
    # bias² decreases as model complexity increases:
    # simple models make systematic errors (high bias)
    # complex models can fit any pattern (low bias)
    # formula 4/x^1.5 gives a smooth decreasing curve

    variance = 0.15 * x**1.8
    # variance increases as model complexity increases:
    # simple models are stable (low variance)
    # complex models are sensitive to training data (high variance)
    # formula 0.15*x^1.8 gives a smooth increasing curve

    total    = bias_sq + variance + 0.5
    # total error = bias² + variance + irreducible noise (0.5 = constant floor)
    # irreducible noise = error that cannot be reduced by any model (inherent in data)

    ax.plot(x, bias_sq,  color='#58A6FF', linewidth=2.5, label='Bias²')
    ax.plot(x, variance, color='#F85149', linewidth=2.5, label='Variance')
    ax.plot(x, total,    color='#3FB950', linewidth=3.0, label='Total Error', linestyle='--')
    ax.axhline(y=0.5, color='#8B949E', linewidth=1.0, linestyle=':', alpha=0.7,
               label='Irreducible Noise')
    # linestyle=':' = dotted line

    # find and mark the minimum total error point (the sweet spot)
    min_idx = np.argmin(total)
    # np.argmin() = returns the INDEX of the smallest value in the array
    ax.scatter([x[min_idx]], [total[min_idx]], color='#FFD700', s=150,
               zorder=5, label=f'Optimal complexity (x={x[min_idx]:.1f})')
    ax.annotate('Sweet Spot\n(min total error)',
                xy=(x[min_idx], total[min_idx]),
                xytext=(x[min_idx] + 1.5, total[min_idx] + 0.8),
                arrowprops=dict(arrowstyle='->', color='#FFD700', lw=1.5),
                color='#FFD700', fontsize=9)

    # shade the underfitting region (left of sweet spot)
    ax.axvspan(0.5, x[min_idx], alpha=0.07, color='#F85149')
    # axvspan(xmin, xmax) = shades a vertical band of the plot
    ax.text(x[min_idx]*0.4, ax.get_ylim()[1]*0.7, 'UNDERFIT\n(High Bias)',
            color='#F85149', fontsize=9, ha='center', alpha=0.9)

    # shade the overfitting region (right of sweet spot)
    ax.axvspan(x[min_idx], 10, alpha=0.07, color='#58A6FF')
    ax.text(x[min_idx]*1.6, ax.get_ylim()[1]*0.7, 'OVERFIT\n(High Variance)',
            color='#58A6FF', fontsize=9, ha='center', alpha=0.9)

    ax.set_title('Bias–Variance Tradeoff', color='#E6EDF3', fontsize=12, pad=12)
    ax.set_xlabel('Model Complexity  →', color='#8B949E', fontsize=10)
    ax.set_ylabel('Error', color='#8B949E', fontsize=10)
    ax.set_ylim(bottom=0)   # y-axis starts at 0 (no negative error)
    ax.legend(facecolor='#21262D', edgecolor='#30363D', labelcolor='#C9D1D9', fontsize=9)

    plt.tight_layout(pad=2.0)
    buf = BytesIO()
    plt.savefig(buf, format='png', dpi=130, bbox_inches='tight',
                facecolor=fig.get_facecolor())
    plt.close(fig)
    return buf


# =============================================================================
# DATA GENERATION + MODEL TRAINING
# Each section builds one complete dataset and trains its model(s)
# =============================================================================

# --- 1. LINEAR REGRESSION DATA -----------------------------------------------

X_house_raw = np.array([
    600, 750, 800, 850, 900, 950, 1000, 1100, 1200,
    1300, 1400, 1500, 1600, 1700, 1800, 1900, 2000, 2200, 2400, 2600
])
# np.array() = creates a 1-dimensional NumPy array from a Python list
# these are house sizes in square feet — our single input feature (X)

noise_house = np.random.normal(loc=0, scale=15, size=len(X_house_raw))
# np.random.normal() = generates random numbers from a Gaussian (bell curve) distribution
# loc=0    = mean of the distribution (centred at 0 → noise averages out)
# scale=15 = standard deviation ($15k spread in prices)
# size     = number of values to generate (one per house)

y_house = 50 + 0.12 * X_house_raw + noise_house
# generate house prices using a simple linear formula + noise:
# base price = $50k
# each sqft adds $0.12k ($120) to the price
# noise adds realistic scatter so it's not a perfect line

X_house = X_house_raw.reshape(-1, 1)
# .reshape(-1, 1) = reshapes the 1D array into a 2D column vector
# -1 means "infer this dimension" → numpy figures out the row count automatically
# scikit-learn requires X to be 2D: shape (n_samples, n_features)
# before reshape: shape (20,)   →   after reshape: shape (20, 1)

X_h_train, X_h_test, y_h_train, y_h_test = train_test_split(
    X_house, y_house, test_size=0.3, random_state=42
)
# train_test_split returns FOUR arrays in this order:
# X_h_train = training features (70% of data = 14 houses)
# X_h_test  = test features     (30% of data =  6 houses)
# y_h_train = training labels   (prices for the 14 training houses)
# y_h_test  = test labels       (prices for the 6 test houses)
# test_size=0.3  = 30% of samples go to test set
# random_state=42 = deterministic shuffle (same split every run)

lr_model = LinearRegression()
# LinearRegression() = creates an untrained model object
# at this point the model has no learned weights — just default parameters

lr_model.fit(X_h_train, y_h_train)
# .fit(X, y) = TRAIN the model: find the best coefficient and intercept
# internally: solves the Normal Equation or uses gradient descent
# to minimise Mean Squared Error: MSE = (1/n) Σ(y - ŷ)²

y_h_pred_train = lr_model.predict(X_h_train)
# .predict(X) = apply learned equation to input X and return predicted y values
# uses: ŷ = β₀ + β₁·x  (intercept + coefficient × feature value)

y_h_pred_test = lr_model.predict(X_h_test)
# same but on test set — these are the "new unseen" predictions

lr_r2_train  = r2_score(y_h_train, y_h_pred_train)
# r2_score(y_true, y_pred) = R² coefficient of determination
# formula: R² = 1 - SS_res/SS_tot
# SS_res = sum of squared residuals (prediction errors)
# SS_tot = total variance in the data
# R²=1.0 = perfect fit, R²=0 = predicts the mean, R²<0 = worse than the mean

lr_r2_test   = r2_score(y_h_test,  y_h_pred_test)
lr_rmse_train = np.sqrt(mean_squared_error(y_h_train, y_h_pred_train))
# mean_squared_error = (1/n)Σ(actual - predicted)²
# np.sqrt() = take the square root to get RMSE (same units as target: $k)

lr_rmse_test  = np.sqrt(mean_squared_error(y_h_test,  y_h_pred_test))
lr_coef       = lr_model.coef_[0]
# .coef_ = learned weight/slope for each feature (attribute with _ = learned after fit)
# [0] = first (and only) coefficient since we have one feature
lr_intercept  = lr_model.intercept_
# .intercept_ = the y-intercept (β₀) of the fitted line

# --- 2. LOGISTIC REGRESSION DATA ---------------------------------------------

n_students = 80   # total number of student samples to generate

study_hours = np.random.uniform(low=1, high=10, size=n_students)
# np.random.uniform() = random values from a uniform distribution
# low=1, high=10 = all values equally likely between 1 and 10 hours
# simulates students who studied between 1 and 10 hours

sleep_hours = np.random.uniform(low=4, high=9, size=n_students)
# students slept between 4 and 9 hours

log_odds_exam = -4 + 0.7 * study_hours + 0.3 * sleep_hours + np.random.normal(0, 0.5, n_students)
# log-odds = ln(p/(1-p)) = linear combination of features
# -4 = negative intercept (failing is the default without study)
# 0.7 * study = study hours have strong positive effect on passing
# 0.3 * sleep = sleep has moderate positive effect
# + noise = real-world noise in exam outcomes

prob_pass = 1 / (1 + np.exp(-log_odds_exam))
# sigmoid function: σ(z) = 1 / (1 + e^(-z))
# converts log-odds to probability: output is always between 0 and 1
# np.exp() = e^x (Euler's number raised to the power of each element)

y_exam = (prob_pass > 0.5).astype(int)
# (prob_pass > 0.5) = boolean array: True where probability exceeds 50%
# .astype(int) = convert True→1 (Pass), False→0 (Fail)

X_exam = np.column_stack([study_hours, sleep_hours])
# np.column_stack() = combines two 1D arrays side by side into a 2D array
# result shape: (80, 2) — 80 students, 2 features each (study, sleep)

X_e_train, X_e_test, y_e_train, y_e_test = train_test_split(
    X_exam, y_exam, test_size=0.25, random_state=42
)
# 25% test split = 60 training students, 20 test students

logr_model = LogisticRegression(max_iter=1000)
# LogisticRegression() = creates logistic regression classifier
# max_iter=1000 = maximum number of optimisation iterations
# default is 100 which sometimes fails to converge (hits the limit early)

logr_model.fit(X_e_train, y_e_train)
# trains the model: finds weights w₁ (study), w₂ (sleep), b (bias)
# minimises binary cross-entropy loss: L = -[y·log(ŷ) + (1-y)·log(1-ŷ)]

y_e_pred_train  = logr_model.predict(X_e_train)
y_e_pred_test   = logr_model.predict(X_e_test)
# .predict() for logistic regression: returns class labels (0 or 1)
# internally: compute sigmoid, then apply 0.5 threshold

y_e_prob_test   = logr_model.predict_proba(X_e_test)[:, 1]
# .predict_proba() = returns probabilities for each class
# shape (n_samples, n_classes) = (20, 2) for binary
# [:, 1] = second column = probability of class 1 (Pass)
# [:, 0] would be probability of class 0 (Fail)

logr_acc_train  = accuracy_score(y_e_train, y_e_pred_train)
# accuracy_score = (number of correct predictions) / (total predictions)
logr_acc_test   = accuracy_score(y_e_test,  y_e_pred_test)
logr_cm         = confusion_matrix(y_e_test, y_e_pred_test)
# confusion_matrix returns a 2×2 array:
# [[TN, FP],   TN = True Negatives  (predicted Fail, actually Fail)
#  [FN, TP]]   FP = False Positives (predicted Pass, actually Fail) ← error
#              FN = False Negatives (predicted Fail, actually Pass) ← error
#              TP = True Positives  (predicted Pass, actually Pass)

# --- 3. UNDERFITTING DATA ----------------------------------------------------

X_poly_raw = np.linspace(-3, 3, 60)
# np.linspace(-3, 3, 60) = 60 evenly spaced values between -3 and +3
# simulates x values on a number line (like positions, times, angles)

y_poly = X_poly_raw**3 - 2 * X_poly_raw**2 + X_poly_raw + np.random.normal(0, 2, 60)
# true underlying pattern is a CUBIC function: y = x³ - 2x² + x
# ** = Python exponentiation operator (3**2 = 9)
# + noise = adds realistic scatter (standard deviation of 2)
# a linear (degree-1) model CANNOT fit this curved pattern → underfitting

X_poly = X_poly_raw.reshape(-1, 1)   # reshape to 2D for sklearn compatibility

X_p_train, X_p_test, y_p_train, y_p_test = train_test_split(
    X_poly, y_poly, test_size=0.3, random_state=42
)

# Underfitting model: degree-1 (just a straight line)
under_model = LinearRegression()
under_model.fit(X_p_train, y_p_train)
# fit() on degree-1 data: finds best-fit straight line
# result: a LINE through curved data = systematic errors everywhere

y_under_pred_train = under_model.predict(X_p_train)
y_under_pred_test  = under_model.predict(X_p_test)

# Good-fit model: degree-3 polynomial
poly3 = PolynomialFeatures(degree=3, include_bias=False)
# PolynomialFeatures(degree=3) = transforms [x] into [x, x², x³]
# include_bias=False = don't add a column of 1s (LinearRegression adds bias itself)
# this gives the linear model the ABILITY to fit cubic curves

X_p_train_poly3 = poly3.fit_transform(X_p_train)
# fit_transform() = fit (learn the polynomial expansion) AND transform (apply it)
# on training data: learn AND apply
# result shape changes: (42, 1) → (42, 3) — three features: x, x², x³

X_p_test_poly3  = poly3.transform(X_p_test)
# .transform() = apply the ALREADY-FITTED expansion (do NOT re-fit on test data)
# critical rule: always fit preprocessing on TRAIN only, apply to test
# avoids data leakage (test data influencing the feature engineering)

good_poly_model = LinearRegression()
good_poly_model.fit(X_p_train_poly3, y_p_train)
# fits a linear model to [x, x², x³] features
# effectively fits: y = β₀ + β₁x + β₂x² + β₃x³
# this can match the true cubic pattern

y_good_pred_train = good_poly_model.predict(X_p_train_poly3)
y_good_pred_test  = good_poly_model.predict(X_p_test_poly3)

uf_r2_under_tr = r2_score(y_p_train, y_under_pred_train)
uf_r2_under_te = r2_score(y_p_test,  y_under_pred_test)
uf_r2_good_tr  = r2_score(y_p_train, y_good_pred_train)
uf_r2_good_te  = r2_score(y_p_test,  y_good_pred_test)
uf_rmse_under_tr = np.sqrt(mean_squared_error(y_p_train, y_under_pred_train))
uf_rmse_under_te = np.sqrt(mean_squared_error(y_p_test,  y_under_pred_test))
uf_rmse_good_tr  = np.sqrt(mean_squared_error(y_p_train, y_good_pred_train))
uf_rmse_good_te  = np.sqrt(mean_squared_error(y_p_test,  y_good_pred_test))

# --- 4. OVERFITTING DATA -----------------------------------------------------

n_emails = 120   # total emails in dataset

word_count   = np.random.randint(low=10,  high=500, size=n_emails)
# np.random.randint() = random INTEGER values in [low, high) — high is excluded
# simulates emails with 10 to 499 words

exclamations = np.random.randint(low=0,   high=20,  size=n_emails)
# number of exclamation marks: 0 to 19

caps_ratio   = np.random.uniform(low=0.0, high=1.0, size=n_emails)
# proportion of CAPITALISED characters: 0.0 (none) to 1.0 (all caps)

spam_log_odds = (-2 + 0.004 * word_count + 0.2 * exclamations + 1.5 * caps_ratio
                 + np.random.normal(0, 0.5, n_emails))
# linear combination defining spam probability:
# long emails with many !!!s and ALL CAPS → higher spam probability

y_spam = (1 / (1 + np.exp(-spam_log_odds)) > 0.5).astype(int)
# same sigmoid + threshold trick as logistic regression data generation

X_spam = np.column_stack([word_count, exclamations, caps_ratio])
# 3-feature matrix: (120, 3)

X_s_train, X_s_test, y_s_train, y_s_test = train_test_split(
    X_spam, y_spam, test_size=0.3, random_state=42
)

# Overfitting model: unlimited depth decision tree
over_tree = DecisionTreeClassifier(max_depth=None, random_state=42)
# max_depth=None = no limit on tree depth
# with enough depth, a decision tree can perfectly memorise every training example
# it creates one leaf node per training point = 100% train accuracy
# but these extremely specific rules don't generalise to new data

over_tree.fit(X_s_train, y_s_train)
# fit() builds the tree by recursively splitting on features
# with no depth limit, splits continue until each leaf is "pure" (one class)

# Good-fit model: depth-4 tree
good_tree = DecisionTreeClassifier(max_depth=4, random_state=42)
# max_depth=4 = tree can ask at most 4 yes/no questions before making a prediction
# forces the model to learn GENERAL rules, not memorise individual emails

good_tree.fit(X_s_train, y_s_train)

of_acc_over_tr = accuracy_score(y_s_train, over_tree.predict(X_s_train))
of_acc_over_te = accuracy_score(y_s_test,  over_tree.predict(X_s_test))
of_acc_good_tr = accuracy_score(y_s_train, good_tree.predict(X_s_train))
of_acc_good_te = accuracy_score(y_s_test,  good_tree.predict(X_s_test))
over_pred_on_train = over_tree.predict(X_s_train)   # for chart
good_pred_on_train = good_tree.predict(X_s_train)   # for chart

# =============================================================================
# GENERATE ALL CHARTS
# =============================================================================

print("Generating charts...")   # print() = outputs text to the terminal/console

buf_lr  = chart_linear_regression(
    X_h_train.flatten(), y_h_train, X_h_test.flatten(), y_h_test,
    y_h_pred_train, y_h_pred_test, lr_coef, lr_intercept
)
# .flatten() = converts 2D column array shape (n,1) back to 1D shape (n,)
# needed because the chart functions expect simple 1D arrays for scatter plots

buf_logr = chart_logistic_regression(
    X_e_test, y_e_test, y_e_pred_test, y_e_prob_test
)

buf_uf   = chart_underfitting(
    X_p_train.flatten(), y_p_train, X_p_test.flatten(), y_p_test,
    y_under_pred_train, y_under_pred_test, y_good_pred_train, y_good_pred_test
)

buf_of   = chart_overfitting(
    of_acc_over_tr, of_acc_over_te, of_acc_good_tr, of_acc_good_te,
    X_s_train, y_s_train, X_s_test, y_s_test,
    over_pred_on_train, good_pred_on_train
)

buf_bv   = chart_bias_variance()   # bias-variance tradeoff diagram

print("All charts generated.")

# =============================================================================
# BUILD EXCEL WORKBOOK
# =============================================================================

wb = Workbook()
# Workbook() = creates a new empty Excel workbook in memory
# to save it to disk we call wb.save('filename.xlsx') at the end

wb.remove(wb.active)
# wb.active = the default sheet Excel creates automatically ("Sheet")
# wb.remove() = deletes that empty sheet so we start completely fresh

# =============================================================================
# SHEET 1: OVERVIEW
# =============================================================================

ws_ov = wb.create_sheet("📋 Overview")
# create_sheet() = adds a new worksheet tab to the workbook
# the string becomes the tab name visible at the bottom of Excel

ws_ov.sheet_view.showGridLines = False
# showGridLines=False = hides the grey grid lines in the background
# gives a cleaner, more polished spreadsheet appearance

set_col_widths(ws_ov, [3, 30, 25, 25, 30, 3])
# column widths in Excel units (approximately characters wide)
# col A=3 (spacer), B=30, C=25, D=25, E=30, F=3 (spacer)

ws_ov.row_dimensions[1].height = 10   # small top padding row

# Title cell
tc = ws_ov.cell(row=2, column=2, value="ML Programs — Detailed Guide with Charts")
tc.font      = Font(name='Arial', bold=True, size=16, color='FFFFFF')
tc.fill      = hex_fill('1A1A2E')
tc.alignment = Alignment(horizontal='left', vertical='center', indent=2)
ws_ov.row_dimensions[2].height = 36
ws_ov.merge_cells('B2:E2')
# merge_cells() = combines B2 through E2 into one wide cell for the title

subtitle = ws_ov.cell(row=3, column=2,
                      value="Programs: Linear Regression | Logistic Regression | Underfitting | Overfitting")
subtitle.font      = Font(name='Arial', italic=True, size=9, color='8B949E')
subtitle.fill      = hex_fill('0D1117')
subtitle.alignment = Alignment(horizontal='left', vertical='center', indent=2)
ws_ov.merge_cells('B3:E3')
ws_ov.row_dimensions[3].height = 18

# Table of contents
style_header(ws_ov, 5, 2, "SHEET", '2C3E7A', colspan=1)
style_header(ws_ov, 5, 3, "ALGORITHM", '2C3E7A', colspan=1)
style_header(ws_ov, 5, 4, "TASK TYPE", '2C3E7A', colspan=1)
style_header(ws_ov, 5, 5, "KEY CONCEPT", '2C3E7A', colspan=1)
ws_ov.row_dimensions[5].height = 20

toc = [
    ("📈 Linear Regression",    "Linear Regression",              "Regression",           "y = β₀ + β₁x  →  minimise MSE  →  R² measures fit quality"),
    ("🎯 Logistic Regression",  "Logistic Regression",            "Binary Classification","σ(z) squashes linear combo to probability  →  threshold at 0.5"),
    ("📉 Underfitting",         "Polynomial (degree 1 vs 3)",     "Regression",           "High bias: too simple a model for a cubic pattern"),
    ("🔥 Overfitting",          "Decision Tree (∞ vs depth-4)",   "Classification",       "High variance: unlimited tree memorises noise, fails on new data"),
    ("📊 Bias-Variance",        "Theoretical framework",          "Core Concept",         "Total Error = Bias² + Variance + Noise — the central ML tradeoff"),
]
row_bg = ["1E2A3A", "16202E", "1E2A3A", "16202E", "1E2A3A"]
for i, (sheet, algo, task, concept) in enumerate(toc):
    r = 6 + i
    bg = row_bg[i]
    for c, val in enumerate([sheet, algo, task, concept], 2):
        cell = ws_ov.cell(row=r, column=c, value=val)
        cell.font      = Font(name='Arial', size=9, color='C9D1D9')
        cell.fill      = hex_fill(bg)
        cell.alignment = Alignment(horizontal='left', vertical='center',
                                   indent=1, wrap_text=True)
        cell.border    = thin_border()
    ws_ov.row_dimensions[r].height = 24

# Bias-Variance chart on overview sheet
insert_chart_image(ws_ov, buf_bv, 'B12', width_px=620, height_px=380)

# =============================================================================
# SHEET 2: LINEAR REGRESSION
# =============================================================================

ws_lr = wb.create_sheet("📈 Linear Regression")
ws_lr.sheet_view.showGridLines = False
set_col_widths(ws_lr, [2, 22, 20, 20, 20, 2, 18, 18, 18, 18, 2])

ws_lr.row_dimensions[1].height = 8

style_header(ws_lr, 2, 2, "LINEAR REGRESSION  —  House Size → Price", '1A1A2E', colspan=9)
ws_lr.row_dimensions[2].height = 28

# explanation rows
explanations = [
    "WHAT:   Fits the straight line  y = β₀ + β₁·x  that minimises sum of squared errors.",
    "DATA:   20 houses. Feature X = size (sqft). Target y = price ($000s).",
    "SPLIT:  70% training (14 houses) / 30% test (6 houses).",
    "OUTPUT: Intercept β₀ (base price) + Coefficient β₁ (price increase per sqft).",
]
for i, txt in enumerate(explanations):
    r = 3 + i
    c = ws_lr.cell(row=r, column=2, value=txt)
    c.font      = Font(name='Arial', italic=True, size=9, color='8B949E')
    c.fill      = hex_fill('0D1117')
    c.alignment = left_indent()
    ws_lr.merge_cells(f'B{r}:J{r}')
    ws_lr.row_dimensions[r].height = 16

# Model metrics section
style_header(ws_lr, 8, 2, "MODEL EQUATION & METRICS", '16502A', colspan=4)
metrics_lr = [
    ("Intercept β₀",         f"${lr_intercept:.2f}k",    "Base price when size = 0 (theoretical starting point)"),
    ("Coefficient β₁",       f"${lr_coef:.4f}k/sqft",    "Each extra sqft adds this much to the predicted price"),
    ("Equation",             f"Price = {lr_intercept:.1f} + {lr_coef:.4f}×Size", "The full learned prediction formula"),
    ("R² — Train",           f"{lr_r2_train:.4f}",        "1.0=perfect; value shows how much variance is explained"),
    ("R² — Test",            f"{lr_r2_test:.4f}",         "Should be close to Train R² (indicates no overfitting)"),
    ("RMSE — Train",         f"${lr_rmse_train:.2f}k",    "Average prediction error on training data"),
    ("RMSE — Test",          f"${lr_rmse_test:.2f}k",     "Average prediction error on unseen test data"),
    ("Overfit?",             "NO — Train ≈ Test R²",       "Small gap between train/test confirms good generalisation"),
]
for i, (key, val, explanation) in enumerate(metrics_lr):
    r = 9 + i
    write_kv(ws_lr, r, 2, key, val)
    # add explanation in a third cell
    ex_cell = ws_lr.cell(row=r, column=4, value=explanation)
    ex_cell.font      = Font(name='Arial', size=8, color='8B949E', italic=True)
    ex_cell.fill      = hex_fill('0D1117')
    ex_cell.alignment = left_indent()
    ex_cell.border    = thin_border()
    ws_lr.merge_cells(f'D{r}:E{r}')

# Training data table
style_header(ws_lr, 8, 7, "TRAINING DATA", '16502A', colspan=4)
train_heads = ["Size (sqft)", "Actual Price", "Predicted", "Residual"]
for c, h in enumerate(train_heads, 7):
    cell = ws_lr.cell(row=9, column=c, value=h)
    cell.font      = hdr_font()
    cell.fill      = hex_fill('16502A')
    cell.alignment = center()
    cell.border    = thin_border()
ws_lr.row_dimensions[9].height = 18

sort_tr = np.argsort(X_h_train.flatten())
# argsort on flattened array gives indices to sort ascending by house size
for i, idx in enumerate(sort_tr):
    r = 10 + i
    vals = [
        int(X_h_train.flatten()[idx]),          # house size as integer (no decimal)
        round(float(y_h_train[idx]), 2),        # actual price rounded to 2 decimals
        round(float(y_h_pred_train[idx]), 2),   # predicted price
        round(float(y_h_train[idx]) - float(y_h_pred_train[idx]), 2)  # residual = actual-pred
    ]
    # float() = explicitly convert numpy float64 to Python float for clean Excel display
    # int()   = explicitly convert numpy int to Python int
    bg = '1A2A1A' if i % 2 == 0 else '0F1F0F'
    # alternating row colours for readability (striped table effect)
    # i % 2 = modulo 2 = remainder when dividing by 2: even rows→0, odd rows→1
    for c, v in enumerate(vals, 7):
        cell = ws_lr.cell(row=r, column=c, value=v)
        cell.font      = body_font(color='C9D1D9', size=9)
        cell.fill      = hex_fill(bg)
        cell.alignment = center()
        cell.border    = thin_border()
    ws_lr.row_dimensions[r].height = 16

# Test data table
test_start_row = 10 + len(sort_tr) + 2
style_header(ws_lr, test_start_row, 7, "TEST DATA (unseen)", '6B2C2C', colspan=4)
for c, h in enumerate(train_heads, 7):
    cell = ws_lr.cell(row=test_start_row+1, column=c, value=h)
    cell.font      = hdr_font()
    cell.fill      = hex_fill('6B2C2C')
    cell.alignment = center()
    cell.border    = thin_border()
ws_lr.row_dimensions[test_start_row+1].height = 18

sort_te = np.argsort(X_h_test.flatten())
for i, idx in enumerate(sort_te):
    r = test_start_row + 2 + i
    vals = [
        int(X_h_test.flatten()[idx]),
        round(float(y_h_test[idx]), 2),
        round(float(y_h_pred_test[idx]), 2),
        round(float(y_h_test[idx]) - float(y_h_pred_test[idx]), 2)
    ]
    bg = '2A1A1A' if i % 2 == 0 else '1F0F0F'
    for c, v in enumerate(vals, 7):
        cell = ws_lr.cell(row=r, column=c, value=v)
        cell.font      = body_font(color='C9D1D9', size=9)
        cell.fill      = hex_fill(bg)
        cell.alignment = center()
        cell.border    = thin_border()
    ws_lr.row_dimensions[r].height = 16

# Embed the linear regression chart
chart_start = test_start_row + len(sort_te) + 3
insert_chart_image(ws_lr, buf_lr, f'B{chart_start}', width_px=700, height_px=360)

# =============================================================================
# SHEET 3: LOGISTIC REGRESSION
# =============================================================================

ws_logr = wb.create_sheet("🎯 Logistic Regression")
ws_logr.sheet_view.showGridLines = False
set_col_widths(ws_logr, [2, 22, 20, 20, 20, 2, 16, 16, 16, 16, 16, 2])

ws_logr.row_dimensions[1].height = 8
style_header(ws_logr, 2, 2, "LOGISTIC REGRESSION  —  Exam Pass/Fail Predictor", '1A1A2E', colspan=10)
ws_logr.row_dimensions[2].height = 28

logr_explanations = [
    "WHAT:   Models P(pass) = σ(β₀ + β₁·study + β₂·sleep) where σ is the sigmoid function.",
    "DATA:   80 students. Features: study hours (1-10) + sleep hours (4-9). Label: pass=1 / fail=0.",
    "SPLIT:  75% training (60 students) / 25% test (20 students).",
    "OUTPUT: Probabilities for each student; threshold at 0.5 gives final Pass/Fail prediction.",
]
for i, txt in enumerate(logr_explanations):
    r = 3 + i
    c = ws_logr.cell(row=r, column=2, value=txt)
    c.font = Font(name='Arial', italic=True, size=9, color='8B949E')
    c.fill = hex_fill('0D1117')
    c.alignment = left_indent()
    ws_logr.merge_cells(f'B{r}:K{r}')
    ws_logr.row_dimensions[r].height = 16

# Metrics
style_header(ws_logr, 8, 2, "MODEL COEFFICIENTS & METRICS", '2C3E7A', colspan=4)
logr_metrics = [
    ("Intercept β₀",    f"{logr_model.intercept_[0]:.4f}",     "Log-odds when study=0 and sleep=0 (baseline)"),
    ("β₁ — Study hrs",  f"{logr_model.coef_[0][0]:.4f}",       "Each study hour increases log-odds of passing by this amount"),
    ("β₂ — Sleep hrs",  f"{logr_model.coef_[0][1]:.4f}",       "Each sleep hour contributes this to the log-odds of passing"),
    ("Train Accuracy",  f"{logr_acc_train*100:.1f}%",           "Fraction of training students correctly classified"),
    ("Test Accuracy",   f"{logr_acc_test*100:.1f}%",            "Fraction of test students correctly classified (key metric)"),
    ("Train-Test Gap",  f"{abs(logr_acc_train-logr_acc_test)*100:.1f}%", "Small gap → no overfitting"),
]
for i, (key, val, explanation) in enumerate(logr_metrics):
    r = 9 + i
    write_kv(ws_logr, r, 2, key, val)
    ex_cell = ws_logr.cell(row=r, column=4, value=explanation)
    ex_cell.font = Font(name='Arial', size=8, color='8B949E', italic=True)
    ex_cell.fill = hex_fill('0D1117')
    ex_cell.alignment = left_indent()
    ex_cell.border = thin_border()
    ws_logr.merge_cells(f'D{r}:E{r}')

# Confusion Matrix
cm_row = 9 + len(logr_metrics) + 2
style_header(ws_logr, cm_row, 2, "CONFUSION MATRIX (Test Set)", '2C3E7A', colspan=4)
ws_logr.row_dimensions[cm_row].height = 20

cm_layout = [
    ["",             "Pred: FAIL",    "Pred: PASS"],
    ["Actual: FAIL", int(logr_cm[0][0]), int(logr_cm[0][1])],
    ["Actual: PASS", int(logr_cm[1][0]), int(logr_cm[1][1])],
]
cm_styles = [
    [('2C3E7A','FFFFFF'), ('2C3E7A','FFFFFF'), ('2C3E7A','FFFFFF')],
    [('2C3E7A','FFFFFF'), ('1A3A1A','3FB950'), ('3A1A1A','F85149')],
    [('2C3E7A','FFFFFF'), ('3A1A1A','F85149'), ('1A3A1A','3FB950')],
]
# diagonal (TN, TP) = green (correct), off-diagonal (FP, FN) = red (errors)
for ri, row_data in enumerate(cm_layout):
    for ci, val in enumerate(row_data):
        cell = ws_logr.cell(row=cm_row+1+ri, column=2+ci, value=val)
        bg, fg = cm_styles[ri][ci]
        cell.font = Font(name='Arial', bold=True, size=10, color=fg)
        cell.fill = hex_fill(bg)
        cell.alignment = center()
        cell.border = thin_border()
    ws_logr.row_dimensions[cm_row+1+ri].height = 24

# Test data table
style_header(ws_logr, 8, 7, "TEST PREDICTIONS (all 20 students)", '5A2D82', colspan=5)
logr_heads = ["Study Hrs", "Sleep Hrs", "Actual", "Predicted", "P(Pass)"]
for c, h in enumerate(logr_heads, 7):
    cell = ws_logr.cell(row=9, column=c, value=h)
    cell.font = hdr_font(); cell.fill = hex_fill('5A2D82')
    cell.alignment = center(); cell.border = thin_border()
ws_logr.row_dimensions[9].height = 18

for i in range(len(y_e_test)):
    r = 10 + i
    correct = bool(y_e_test[i] == y_e_pred_test[i])
    # bool() = convert numpy bool_ to Python bool for reliable comparisons
    vals = [
        round(float(X_e_test[i][0]), 2),   # study hours
        round(float(X_e_test[i][1]), 2),   # sleep hours
        "Pass" if y_e_test[i] == 1 else "Fail",      # actual label
        "Pass" if y_e_pred_test[i] == 1 else "Fail", # predicted label
        round(float(y_e_prob_test[i]), 3),            # probability of passing
    ]
    col_colors = ['131320','131320','1A3A1A' if y_e_test[i]==1 else '3A1A1A',
                  '1A2A1A' if correct else '3A1A1A',
                  '1A3A1A' if y_e_prob_test[i]>0.6 else ('3A1A1A' if y_e_prob_test[i]<0.4 else '2A2A10')]
    text_colors = ['C9D1D9','C9D1D9','3FB950' if y_e_test[i]==1 else 'F85149',
                   '3FB950' if correct else 'F85149',
                   '3FB950' if y_e_prob_test[i]>0.6 else ('F85149' if y_e_prob_test[i]<0.4 else 'FFD700')]
    for c, (v, bg, fg) in enumerate(zip(vals, col_colors, text_colors), 7):
        cell = ws_logr.cell(row=r, column=c, value=v)
        cell.font = Font(name='Arial', bold=(c==9 or c==10), size=9, color=fg)
        cell.fill = hex_fill(bg)
        cell.alignment = center(); cell.border = thin_border()
    ws_logr.row_dimensions[r].height = 16

# Embed logistic regression chart
logr_chart_row = 10 + len(y_e_test) + 2
insert_chart_image(ws_logr, buf_logr, f'B{logr_chart_row}', width_px=700, height_px=360)

# =============================================================================
# SHEET 4: UNDERFITTING
# =============================================================================

ws_uf = wb.create_sheet("📉 Underfitting")
ws_uf.sheet_view.showGridLines = False
set_col_widths(ws_uf, [2, 24, 18, 18, 18, 18, 2, 16, 16, 16, 16, 2])

ws_uf.row_dimensions[1].height = 8
style_header(ws_uf, 2, 2, "UNDERFITTING  —  Degree-1 Line on Cubic Data", '1A1A2E', colspan=10)
ws_uf.row_dimensions[2].height = 28

uf_expl = [
    "WHAT:   Underfitting occurs when the model is too simple to capture the true data pattern.",
    "DATA:   60 points sampled from y = x³ − 2x² + x + noise. True shape is a CUBIC curve.",
    "UNDERFIT: A straight line (degree-1 polynomial) cannot bend → high error everywhere.",
    "GOOD FIT: Degree-3 polynomial can match the cubic curve → low error on both sets.",
]
for i, txt in enumerate(uf_expl):
    r = 3 + i
    c = ws_uf.cell(row=r, column=2, value=txt)
    c.font = Font(name='Arial', italic=True, size=9, color='8B949E')
    c.fill = hex_fill('0D1117'); c.alignment = left_indent()
    ws_uf.merge_cells(f'B{r}:K{r}'); ws_uf.row_dimensions[r].height = 16

# Model comparison table
style_header(ws_uf, 8, 2, "MODEL COMPARISON", '2C3E7A', colspan=5)
comp_heads = ["Metric", "Underfit (degree 1)", "Good Fit (degree 3)", "Difference", "Interpretation"]
for c, h in enumerate(comp_heads, 2):
    cell = ws_uf.cell(row=9, column=c, value=h)
    cell.font = hdr_font(); cell.fill = hex_fill('2C3E7A')
    cell.alignment = center(); cell.border = thin_border()
ws_uf.row_dimensions[9].height = 18

comp_rows = [
    ("R² Train",    uf_r2_under_tr, uf_r2_good_tr),
    ("R² Test",     uf_r2_under_te, uf_r2_good_te),
    ("RMSE Train",  uf_rmse_under_tr, uf_rmse_good_tr),
    ("RMSE Test",   uf_rmse_under_te, uf_rmse_good_te),
]
interps = [
    "Both low → model cannot learn the pattern (HIGH BIAS)",
    "Both low → confirms generalisation is also bad",
    "Both high → large errors on training data",
    "Both high → large errors on new data too",
]
for i, (metric, under_v, good_v) in enumerate(comp_rows):
    r = 10 + i
    is_rmse = 'RMSE' in metric
    # for RMSE: lower is better; for R²: higher is better
    diff_val = round(good_v - under_v, 4) if not is_rmse else round(under_v - good_v, 4)
    # positive diff = good-fit model is better (by how much)

    row_vals = [
        metric,
        round(under_v, 4),   # underfit model value
        round(good_v,  4),   # good-fit model value
        f"+{diff_val:.4f}" if diff_val >= 0 else f"{diff_val:.4f}",  # formatted difference
        interps[i],
    ]
    bgs = ['1E2A3A', '2A1A1A', '1A2A1A', '1A1A2A', '0D1117']
    fgs = ['C9D1D9', 'F85149', '3FB950', 'FFD700', '8B949E']
    for c, (v, bg, fg) in enumerate(zip(row_vals, bgs, fgs), 2):
        cell = ws_uf.cell(row=r, column=c, value=v)
        cell.font = Font(name='Arial', bold=(c==2), size=9, color=fg)
        cell.fill = hex_fill(bg)
        cell.alignment = Alignment(horizontal='center' if c < 6 else 'left',
                                   vertical='center', indent=1, wrap_text=True)
        cell.border = thin_border()
    ws_uf.row_dimensions[r].height = 20

# Cause and fix box
cause_row = 15
style_header(ws_uf, cause_row, 2, "CAUSES & FIXES", 'B7410E', colspan=5)
causes = [
    ("⚠ Cause", "Model complexity (degree) is lower than the true data complexity", 'F85149'),
    ("⚠ Cause", "Both TRAIN and TEST errors are HIGH — cannot even learn training data", 'F85149'),
    ("⚠ Cause", "More data will NOT help — problem is model architecture, not data size", 'F85149'),
    ("✓ Fix",   "Increase polynomial degree (1→3→5) until train error becomes acceptable", '3FB950'),
    ("✓ Fix",   "Try a more expressive model family: neural network, decision tree, SVR", '3FB950'),
    ("✓ Fix",   "Reduce regularisation strength if L1/L2 penalty is too strong", '3FB950'),
]
for i, (label, desc, fg) in enumerate(causes):
    r = cause_row + 1 + i
    lc = ws_uf.cell(row=r, column=2, value=label)
    lc.font = Font(name='Arial', bold=True, size=9, color=fg)
    lc.fill = hex_fill('1E1A10' if '⚠' in label else '0F1E10')
    lc.alignment = center(); lc.border = thin_border()

    dc = ws_uf.cell(row=r, column=3, value=desc)
    dc.font = Font(name='Arial', size=9, color='C9D1D9')
    dc.fill = hex_fill('141414'); dc.alignment = left_indent()
    dc.border = thin_border()
    ws_uf.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws_uf.row_dimensions[r].height = 18

# Sample data
data_row = cause_row + len(causes) + 2
style_header(ws_uf, data_row, 2, "SAMPLE DATA — First 20 Training Points", '16502A', colspan=5)
uf_heads = ["x", "y actual", "Underfit pred", "Good-fit pred", "Underfit error"]
for c, h in enumerate(uf_heads, 2):
    cell = ws_uf.cell(row=data_row+1, column=c, value=h)
    cell.font = hdr_font(); cell.fill = hex_fill('16502A')
    cell.alignment = center(); cell.border = thin_border()
ws_uf.row_dimensions[data_row+1].height = 18

sort_p = np.argsort(X_p_train.flatten())
# sort training points by x value for cleaner display
for i, idx in enumerate(sort_p[:20]):   # [:20] = slice: first 20 elements only
    r = data_row + 2 + i
    xv     = round(float(X_p_train.flatten()[idx]), 3)
    yv     = round(float(y_p_train[idx]), 3)
    u_pred = round(float(y_under_pred_train[idx]), 3)
    g_pred = round(float(y_good_pred_train[idx]), 3)
    err    = round(abs(yv - u_pred), 3)   # abs() = absolute value (always positive)
    for c, v in enumerate([xv, yv, u_pred, g_pred, err], 2):
        cell = ws_uf.cell(row=r, column=c, value=v)
        cell.font = body_font(color='C9D1D9', size=9)
        cell.fill = hex_fill('1A2A1A' if i%2==0 else '0F1A0F')
        cell.alignment = center(); cell.border = thin_border()
    ws_uf.row_dimensions[r].height = 16

# Embed underfitting chart
uf_chart_row = data_row + 23
insert_chart_image(ws_uf, buf_uf, f'B{uf_chart_row}', width_px=700, height_px=360)

# =============================================================================
# SHEET 5: OVERFITTING
# =============================================================================

ws_of = wb.create_sheet("🔥 Overfitting")
ws_of.sheet_view.showGridLines = False
set_col_widths(ws_of, [2, 26, 20, 20, 20, 2, 16, 16, 16, 16, 2])

ws_of.row_dimensions[1].height = 8
style_header(ws_of, 2, 2, "OVERFITTING  —  Decision Tree Spam Classifier", '1A1A2E', colspan=9)
ws_of.row_dimensions[2].height = 28

of_expl = [
    "WHAT:   Overfitting is when the model learns the training data TOO well — including its noise.",
    "DATA:   120 emails. Features: word count, exclamation marks, caps ratio. Label: spam=1 / not=0.",
    "OVERFIT: Unlimited tree depth → 100% train accuracy by memorising every example.",
    "GOOD FIT: Max depth 4 → learns general rules → similar accuracy on train AND test.",
]
for i, txt in enumerate(of_expl):
    r = 3 + i
    c = ws_of.cell(row=r, column=2, value=txt)
    c.font = Font(name='Arial', italic=True, size=9, color='8B949E')
    c.fill = hex_fill('0D1117'); c.alignment = left_indent()
    ws_of.merge_cells(f'B{r}:J{r}'); ws_of.row_dimensions[r].height = 16

# Model comparison
style_header(ws_of, 8, 2, "ACCURACY COMPARISON", '2C3E7A', colspan=5)
of_comp_heads = ["Metric", "Overfit Tree (∞)", "Good Tree (4)", "Gap", "Verdict"]
for c, h in enumerate(of_comp_heads, 2):
    cell = ws_of.cell(row=9, column=c, value=h)
    cell.font = hdr_font(); cell.fill = hex_fill('2C3E7A')
    cell.alignment = center(); cell.border = thin_border()
ws_of.row_dimensions[9].height = 18

of_rows = [
    ("Train Accuracy",  of_acc_over_tr,  of_acc_good_tr,
     "Overfit=100% is suspicious — indicates memorisation"),
    ("Test Accuracy",   of_acc_over_te,  of_acc_good_te,
     "Large drop for overfit tree = fails on unseen data"),
    ("Train−Test Gap",  of_acc_over_tr - of_acc_over_te,  of_acc_good_tr - of_acc_good_te,
     "Small gap in good tree = true generalisation"),
    ("Tree Depth",      None, None, "∞ = memorises every point; 4 = learns patterns"),
    ("Diagnosis",       None, None, "Overfit: Low Bias + HIGH Variance. Good: balanced."),
]
for i, row_data in enumerate(of_rows):
    r = 10 + i
    metric, over_v, good_v, interp = row_data
    if over_v is not None:
        over_str = f"{over_v*100:.1f}%"
        good_str = f"{good_v*100:.1f}%"
        gap_str  = f"{abs(over_v-good_v)*100:.1f}% diff"
    elif i == 3:
        over_str, good_str, gap_str = "Unlimited", "4", "∞ vs 4"
    else:
        over_str, good_str, gap_str = "High Variance", "Balanced", "Overfit detected"

    row_vals = [metric, over_str, good_str, gap_str, interp]
    of_bgs = ['1E2A3A','2A1A1A','1A2A1A','1A1A2A','0D1117']
    of_fgs = ['C9D1D9','F85149','3FB950','FFD700','8B949E']
    for c, (v, bg, fg) in enumerate(zip(row_vals, of_bgs, of_fgs), 2):
        cell = ws_of.cell(row=r, column=c, value=v)
        cell.font = Font(name='Arial', bold=(c==2), size=9, color=fg)
        cell.fill = hex_fill(bg)
        cell.alignment = Alignment(horizontal='center' if c < 6 else 'left',
                                   vertical='center', indent=1, wrap_text=True)
        cell.border = thin_border()
    ws_of.row_dimensions[r].height = 20

# Symptoms and fixes
sym_row = 16
style_header(ws_of, sym_row, 2, "SYMPTOMS & FIXES", '721C24', colspan=5)
symptoms = [
    ("⚠ Symptom", "Train accuracy = 100% (or very close) — model memorised training set", 'F85149'),
    ("⚠ Symptom", "Large Train-Test accuracy gap (e.g. 100% train vs 72% test)", 'F85149'),
    ("⚠ Symptom", "Extremely complex model: 1000-node tree, 50-layer network", 'F85149'),
    ("✓ Fix",     "Limit tree depth (max_depth=4) — forces general rules over specific ones", '3FB950'),
    ("✓ Fix",     "Collect more training data — harder to memorise 10,000 examples", '3FB950'),
    ("✓ Fix",     "Apply L2 regularisation — penalise large weights, smooths the model", '3FB950'),
    ("✓ Fix",     "Dropout (neural nets) — randomly disable neurons during training", '3FB950'),
    ("✓ Fix",     "Early stopping — halt training when validation loss starts rising", '3FB950'),
]
for i, (label, desc, fg) in enumerate(symptoms):
    r = sym_row + 1 + i
    lc = ws_of.cell(row=r, column=2, value=label)
    lc.font = Font(name='Arial', bold=True, size=9, color=fg)
    lc.fill = hex_fill('1E1010' if '⚠' in label else '0F1E10')
    lc.alignment = center(); lc.border = thin_border()
    dc = ws_of.cell(row=r, column=3, value=desc)
    dc.font = Font(name='Arial', size=9, color='C9D1D9')
    dc.fill = hex_fill('141414'); dc.alignment = left_indent()
    dc.border = thin_border()
    ws_of.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    ws_of.row_dimensions[r].height = 18

# Sample data
of_data_row = sym_row + len(symptoms) + 2
style_header(ws_of, of_data_row, 2, "SAMPLE TRAINING DATA (first 20)", 'B7410E', colspan=5)
of_heads = ["Word Count", "Exclamations", "Caps Ratio", "Label", "Spam?"]
for c, h in enumerate(of_heads, 2):
    cell = ws_of.cell(row=of_data_row+1, column=c, value=h)
    cell.font = hdr_font(); cell.fill = hex_fill('B7410E')
    cell.alignment = center(); cell.border = thin_border()
ws_of.row_dimensions[of_data_row+1].height = 18

for i in range(min(20, len(X_s_train))):
    r = of_data_row + 2 + i
    is_spam = int(y_s_train[i]) == 1
    vals = [
        int(X_s_train[i][0]),                            # word count
        int(X_s_train[i][1]),                            # exclamation count
        round(float(X_s_train[i][2]), 3),                # caps ratio
        int(y_s_train[i]),                               # raw label (0 or 1)
        "SPAM" if is_spam else "Not Spam",               # readable label
    ]
    bgs = ['2A1010' if is_spam else '0F1A0F',
           '2A1010' if is_spam else '0F1A0F',
           '2A1010' if is_spam else '0F1A0F',
           '2A1010' if is_spam else '0F1A0F',
           '3A0000' if is_spam else '003A00']
    fgs = ['C9D1D9','C9D1D9','C9D1D9','FFD700',
           'F85149' if is_spam else '3FB950']
    for c, (v, bg, fg) in enumerate(zip(vals, bgs, fgs), 2):
        cell = ws_of.cell(row=r, column=c, value=v)
        cell.font = Font(name='Arial', bold=(c==6), size=9, color=fg)
        cell.fill = hex_fill(bg)
        cell.alignment = center(); cell.border = thin_border()
    ws_of.row_dimensions[r].height = 16

# Embed overfitting chart
of_chart_row = of_data_row + 23
insert_chart_image(ws_of, buf_of, f'B{of_chart_row}', width_px=700, height_px=360)

# =============================================================================
# SHEET 6: PYTHON CODE WITH LINE-BY-LINE COMMENTS
# =============================================================================

ws_code = wb.create_sheet("🐍 Annotated Code")
ws_code.sheet_view.showGridLines = False
set_col_widths(ws_code, [2, 55, 55, 2])
# col B = code, col C = comment/explanation

ws_code.row_dimensions[1].height = 8
style_header(ws_code, 2, 2, "ANNOTATED PYTHON CODE — Every Line Explained", '1A1A2E', colspan=2)
ws_code.row_dimensions[2].height = 28

sub = ws_code.cell(row=3, column=2,
                   value="Column B = Python code line    |    Column C = what each keyword/function does")
sub.font = Font(name='Arial', italic=True, size=9, color='8B949E')
sub.fill = hex_fill('0D1117'); sub.alignment = left_indent()
ws_code.merge_cells('B3:C3')
ws_code.row_dimensions[3].height = 16

# Each tuple: (code_line, explanation, section_color or None)
# section_color = hex string → it's a section header row; None = regular code line
annotated_lines = [
    # ── Linear Regression block ──────────────────────────────────────────────
    ("# ══ LINEAR REGRESSION ══", None, '16502A'),
    ("import numpy as np",
     "import = load a library. numpy = numerical computing. np = short alias so we write np.array() instead of numpy.array()", None),
    ("from sklearn.linear_model import LinearRegression",
     "from...import = load only one class from a library. LinearRegression finds the best-fit straight line through data.", None),
    ("from sklearn.model_selection import train_test_split",
     "train_test_split = splits data into train set (model learns from this) and test set (evaluate on unseen data).", None),
    ("from sklearn.metrics import r2_score, mean_squared_error",
     "r2_score = R² coefficient (1=perfect fit). mean_squared_error = average of (actual-predicted)².", None),
    ("",  "", None),
    ("np.random.seed(42)",
     "seed() = fix the random number generator to a specific starting value. Ensures SAME results every run (reproducibility).", None),
    ("X = np.array([600,750,...]).reshape(-1,1)",
     "np.array() = create a NumPy array. reshape(-1,1) = make it 2D (n rows × 1 column). sklearn needs 2D input.", None),
    ("y = 50 + 0.12 * X.flatten() + np.random.normal(0,15,20)",
     "Generate prices. 50=base price. 0.12=price per sqft. normal(mean,std,n) = add Gaussian noise for realism.", None),
    ("X_train,X_test,y_train,y_test = train_test_split(X,y,test_size=0.3,random_state=42)",
     "Split: test_size=0.3 means 30% goes to test. random_state=42 = reproducible shuffle. Returns 4 arrays.", None),
    ("model = LinearRegression()",
     "Create an UNTRAINED model object. Parentheses call the constructor. Model has no learned values yet.", None),
    ("model.fit(X_train, y_train)",
     ".fit() = TRAIN the model. Finds best β₀ (intercept) and β₁ (coefficient) by minimising MSE.", None),
    ("y_pred = model.predict(X_test)",
     ".predict() = apply learned equation ŷ = β₀ + β₁·x to new input X. Returns predicted values.", None),
    ("r2 = r2_score(y_test, y_pred)",
     "r2_score(actual, predicted) = R² = 1 - SS_res/SS_tot. Range: −∞ to 1.0. Higher = better fit.", None),
    ("rmse = np.sqrt(mean_squared_error(y_test, y_pred))",
     "np.sqrt() = square root. MSE = mean of (actual−pred)². RMSE = √MSE = error in same units as target.", None),
    ("coef = model.coef_[0]",
     "model.coef_ = learned slope (attribute set AFTER fit, hence trailing underscore). [0] = first feature.", None),
    ("intercept = model.intercept_",
     "model.intercept_ = learned y-intercept β₀. Together: equation is y = intercept + coef * x", None),

    # ── Logistic Regression block ─────────────────────────────────────────────
    ("", "", None),
    ("# ══ LOGISTIC REGRESSION ══", None, '2C3E7A'),
    ("from sklearn.linear_model import LogisticRegression",
     "LogisticRegression = CLASSIFICATION model (despite name). Uses sigmoid to output probabilities.", None),
    ("from sklearn.metrics import accuracy_score, confusion_matrix",
     "accuracy_score = correct/total. confusion_matrix = table of TP,TN,FP,FN counts.", None),
    ("",  "", None),
    ("study = np.random.uniform(1, 10, 80)",
     "uniform(low, high, size) = random floats with EQUAL probability between low and high. 80 student samples.", None),
    ("log_odds = -4 + 0.7*study + 0.3*sleep",
     "Log-odds = linear combination. -4=baseline (hard to pass without study). 0.7=effect of each study hour.", None),
    ("prob = 1 / (1 + np.exp(-log_odds))",
     "Sigmoid function σ(z) = 1/(1+e^-z). np.exp() = e^x. Converts log-odds to probability (always 0 to 1).", None),
    ("y = (prob > 0.5).astype(int)",
     "(prob>0.5) = boolean array True/False. .astype(int) converts True→1 (Pass), False→0 (Fail).", None),
    ("X = np.column_stack([study, sleep])",
     "column_stack = combine two 1D arrays into a 2D matrix side by side. Result: (80,2) array.", None),
    ("model = LogisticRegression(max_iter=1000)",
     "max_iter=1000 = allow up to 1000 optimization steps. Default 100 sometimes fails to converge.", None),
    ("model.fit(X_train, y_train)",
     "Trains by minimising binary cross-entropy: L = -[y·log(ŷ) + (1-y)·log(1-ŷ)]", None),
    ("probs = model.predict_proba(X_test)[:, 1]",
     "predict_proba() = returns (n,2) array: col0=P(Fail), col1=P(Pass). [:, 1] selects P(Pass) column.", None),
    ("cm = confusion_matrix(y_test, y_pred)",
     "confusion_matrix returns [[TN,FP],[FN,TP]]. Diagonal=correct, off-diagonal=errors.", None),

    # ── Underfitting block ────────────────────────────────────────────────────
    ("", "", None),
    ("# ══ UNDERFITTING ══", None, 'B7410E'),
    ("from sklearn.preprocessing import PolynomialFeatures",
     "PolynomialFeatures = transforms [x] into [x, x², x³,...]. Gives linear model ability to fit curves.", None),
    ("",  "", None),
    ("X = np.linspace(-3, 3, 60).reshape(-1,1)",
     "linspace(start, stop, n) = n evenly spaced values. reshape(-1,1) = make 2D for sklearn.", None),
    ("y = X**3 - 2*X**2 + X + noise",
     "** = exponentiation (3**2=9). True pattern is cubic. The underfit model won't know this.", None),
    ("under_model = LinearRegression().fit(X_train, y_train)",
     "Chain .fit() directly on constructor. Fits a straight line — cannot capture cubic shape = UNDERFIT.", None),
    ("poly3 = PolynomialFeatures(degree=3, include_bias=False)",
     "degree=3 = create features [x, x², x³]. include_bias=False = don't add 1s column (LinearRegression does it).", None),
    ("X_train_p3 = poly3.fit_transform(X_train)",
     "fit_transform = fit (on train only) then transform. Rule: NEVER fit on test data. Shape: (n,1)→(n,3).", None),
    ("X_test_p3 = poly3.transform(X_test)",
     "transform only (NOT fit_transform) on test set. Uses parameters learned from training data only.", None),
    ("good_model = LinearRegression().fit(X_train_p3, y_train)",
     "Fits β₀+β₁x+β₂x²+β₃x³ — a cubic equation. Can match the true cubic pattern.", None),

    # ── Overfitting block ─────────────────────────────────────────────────────
    ("", "", None),
    ("# ══ OVERFITTING ══", None, '5A2D82'),
    ("from sklearn.tree import DecisionTreeClassifier",
     "DecisionTreeClassifier = builds a tree of if/else splits. Unlimited depth = memorise every training point.", None),
    ("",  "", None),
    ("word_count = np.random.randint(10, 500, 120)",
     "randint(low, high, size) = random INTEGERS in [low, high). 120 email samples.", None),
    ("y = (sigmoid(spam_log_odds) > 0.5).astype(int)",
     "Generate ground truth spam labels using sigmoid threshold. 1=Spam, 0=Not Spam.", None),
    ("over_tree = DecisionTreeClassifier(max_depth=None)",
     "max_depth=None = no limit. Tree grows until every leaf is pure (100% train accuracy = MEMORISED).", None),
    ("over_tree.fit(X_train, y_train)",
     "Builds deeply branching tree. Each leaf may represent just 1-2 training examples.", None),
    ("good_tree = DecisionTreeClassifier(max_depth=4)",
     "max_depth=4 = at most 4 splits. Forces GENERAL rules. Cannot memorise individual emails.", None),
    ("acc_train = accuracy_score(y_train, over_tree.predict(X_train))",
     "Overfit tree: likely 100% train accuracy. Good tree: similar train and TEST accuracy.", None),
    ("acc_test = accuracy_score(y_test, over_tree.predict(X_test))",
     "Key insight: overfit tree's test accuracy DROPS significantly — it memorised noise, not pattern.", None),
]

code_row = 5
for item in annotated_lines:
    code_text, comment_text, section_color = item

    if section_color is not None:
        # section header row
        style_header(ws_code, code_row, 2, f"  {code_text}", section_color, colspan=2)
        ws_code.row_dimensions[code_row].height = 22
    else:
        # regular code line
        code_cell = ws_code.cell(row=code_row, column=2, value=code_text)
        code_cell.font = Font(name='Courier New', size=8.5,
                              color='79C0FF' if code_text.startswith('from') or
                                                code_text.startswith('import') else
                              ('F85149' if code_text.startswith('#') else 'E6EDF3'))
        # colour keywords differently:
        # blue  = import statements
        # red   = comment lines (starting with #)
        # white = everything else
        code_cell.fill = hex_fill('0D1117')
        code_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        code_cell.border = thin_border()
        ws_code.row_dimensions[code_row].height = 16

        if comment_text:
            comm_cell = ws_code.cell(row=code_row, column=3, value=f"← {comment_text}")
            comm_cell.font = Font(name='Arial', size=8, color='8B949E', italic=True)
            comm_cell.fill = hex_fill('111119')
            comm_cell.alignment = Alignment(horizontal='left', vertical='center',
                                            indent=1, wrap_text=True)
            comm_cell.border = thin_border()
            ws_code.row_dimensions[code_row].height = 28
            # taller rows for lines with long explanations

    code_row += 1   # advance to the next row for the next line

# =============================================================================
# SAVE THE WORKBOOK
# =============================================================================

import os
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ML_Programs_Annotated.xlsx')
# saves the Excel file in the same directory as this script

wb.save(output_path)
# .save(path) = serialises the workbook to disk as a .xlsx file
# after this line, the file is written and can be opened in Excel

print(f"✓ Saved: {output_path}")
# f-string (f"...") = formatted string — variables inside {} are interpolated
# e.g. f"Result: {1+1}" → "Result: 2"