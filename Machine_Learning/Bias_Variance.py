import numpy as np
import warnings
warnings.filterwarnings('ignore')

from sklearn.linear_model import LinearRegression, LogisticRegression
from sklearn.preprocessing import PolynomialFeatures
from sklearn.tree import DecisionTreeClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error, r2_score, accuracy_score, confusion_matrix
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                             GradientFill)
from openpyxl.utils import get_column_letter
import random

np.random.seed(42)
random.seed(42)

# ─── STYLE HELPERS ────────────────────────────────────────────────────────────

def hex_fill(hex_color):
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(bold=True, color="FFFFFF", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def body_font(bold=False, color="1A1A2E", size=9):
    return Font(name="Arial", bold=bold, color=color, size=size)

def style_header_row(ws, row, cols, bg="1A1A2E", fg="FFFFFF"):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = hex_fill(bg)
        cell.font = header_font(color=fg)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border()

def style_data_row(ws, row, cols, even=True):
    bg = "F4F6FB" if even else "FFFFFF"
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = hex_fill(bg)
        cell.font = body_font()
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border()

def write_section_title(ws, row, col, text, color="1A1A2E", colspan=1):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell.fill = hex_fill(color)
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 22
    return cell

def write_kv(ws, row, col, key, val, key_color="E8EAF6", val_color="FFFFFF"):
    k = ws.cell(row=row, column=col, value=key)
    k.font = Font(name="Arial", bold=True, size=9, color="1A1A2E")
    k.fill = hex_fill(key_color)
    k.alignment = Alignment(horizontal='left', indent=1)
    k.border = thin_border()
    v = ws.cell(row=row, column=col+1, value=val)
    v.font = Font(name="Arial", size=9, color="1A1A2E")
    v.fill = hex_fill(val_color)
    v.alignment = Alignment(horizontal='center')
    v.border = thin_border()

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def pill_cell(ws, row, col, text, good=True):
    cell = ws.cell(row=row, column=col, value=text)
    color = "D4EDDA" if good else "F8D7DA"
    font_color = "155724" if good else "721C24"
    cell.fill = hex_fill(color)
    cell.font = Font(name="Arial", bold=True, size=9, color=font_color)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border()

# ─── 1. LINEAR REGRESSION ─────────────────────────────────────────────────────

def build_linear_regression():
    # House size (sqft) → price ($k)
    X_raw = np.array([600,750,800,850,900,950,1000,1100,1200,1300,
                      1400,1500,1600,1700,1800,1900,2000,2200,2400,2600])
    noise = np.random.normal(0, 15, len(X_raw))
    y = 50 + 0.12 * X_raw + noise

    X = X_raw.reshape(-1, 1)
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=42)

    model = LinearRegression()
    model.fit(X_train, y_train)

    y_pred_train = model.predict(X_train)
    y_pred_test  = model.predict(X_test)

    return {
        "X_raw": X_raw, "y": y,
        "X_train": X_train.flatten(), "y_train": y_train,
        "X_test":  X_test.flatten(),  "y_test":  y_test,
        "y_pred_train": y_pred_train,  "y_pred_test": y_pred_test,
        "coef": model.coef_[0], "intercept": model.intercept_,
        "r2_train": r2_score(y_train, y_pred_train),
        "r2_test":  r2_score(y_test,  y_pred_test),
        "rmse_train": np.sqrt(mean_squared_error(y_train, y_pred_train)),
        "rmse_test":  np.sqrt(mean_squared_error(y_test,  y_pred_test)),
    }

# ─── 2. LOGISTIC REGRESSION ───────────────────────────────────────────────────

def build_logistic_regression():
    # Study hours + sleep hours → exam pass/fail
    n = 80
    study = np.random.uniform(1, 10, n)
    sleep = np.random.uniform(4, 9, n)
    log_odds = -4 + 0.7*study + 0.3*sleep + np.random.normal(0, 0.5, n)
    y = (1 / (1 + np.exp(-log_odds)) > 0.5).astype(int)

    X = np.column_stack([study, sleep])
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.25, random_state=42)

    model = LogisticRegression(max_iter=1000)
    model.fit(X_train, y_train)

    y_pred_train = model.predict(X_train)
    y_pred_test  = model.predict(X_test)
    y_prob_test  = model.predict_proba(X_test)[:, 1]

    cm = confusion_matrix(y_test, y_pred_test)

    return {
        "study": study, "sleep": sleep, "y": y,
        "X_train": X_train, "y_train": y_train,
        "X_test": X_test,   "y_test": y_test,
        "y_pred_train": y_pred_train, "y_pred_test": y_pred_test,
        "y_prob_test": y_prob_test,
        "acc_train": accuracy_score(y_train, y_pred_train),
        "acc_test":  accuracy_score(y_test,  y_pred_test),
        "coef_study": model.coef_[0][0],
        "coef_sleep": model.coef_[0][1],
        "intercept":  model.intercept_[0],
        "cm": cm,
    }

# ─── 3. UNDERFITTING ──────────────────────────────────────────────────────────

def build_underfitting():
    # True pattern: cubic. Underfit with linear.
    X_raw = np.linspace(-3, 3, 60)
    y = X_raw**3 - 2*X_raw**2 + X_raw + np.random.normal(0, 2, 60)

    X = X_raw.reshape(-1, 1)
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=42)

    # Underfit: degree-1
    model_under = LinearRegression()
    model_under.fit(X_train, y_train)
    y_pred_under_train = model_under.predict(X_train)
    y_pred_under_test  = model_under.predict(X_test)

    # Good fit: degree-3
    poly3 = PolynomialFeatures(degree=3)
    X_train_p3 = poly3.fit_transform(X_train)
    X_test_p3  = poly3.transform(X_test)
    model_good = LinearRegression()
    model_good.fit(X_train_p3, y_train)
    y_pred_good_train = model_good.predict(X_train_p3)
    y_pred_good_test  = model_good.predict(X_test_p3)

    return {
        "X_raw": X_raw, "y": y,
        "X_train": X_train.flatten(), "y_train": y_train,
        "X_test":  X_test.flatten(),  "y_test":  y_test,
        "y_pred_under_train": y_pred_under_train,
        "y_pred_under_test":  y_pred_under_test,
        "y_pred_good_train":  y_pred_good_train,
        "y_pred_good_test":   y_pred_good_test,
        "r2_under_train": r2_score(y_train, y_pred_under_train),
        "r2_under_test":  r2_score(y_test,  y_pred_under_test),
        "r2_good_train":  r2_score(y_train, y_pred_good_train),
        "r2_good_test":   r2_score(y_test,  y_pred_good_test),
        "rmse_under_train": np.sqrt(mean_squared_error(y_train, y_pred_under_train)),
        "rmse_under_test":  np.sqrt(mean_squared_error(y_test,  y_pred_under_test)),
        "rmse_good_train":  np.sqrt(mean_squared_error(y_train, y_pred_good_train)),
        "rmse_good_test":   np.sqrt(mean_squared_error(y_test,  y_pred_good_test)),
    }

# ─── 4. OVERFITTING ───────────────────────────────────────────────────────────

def build_overfitting():
    # Classification: spam detection. Overfit with max_depth=None tree.
    n = 120
    word_count    = np.random.randint(10, 500, n)
    exclamations  = np.random.randint(0, 20, n)
    caps_ratio    = np.random.uniform(0, 1, n)

    log_odds = -2 + 0.004*word_count + 0.2*exclamations + 1.5*caps_ratio
    y = (1/(1+np.exp(-log_odds + np.random.normal(0,0.5,n))) > 0.5).astype(int)

    X = np.column_stack([word_count, exclamations, caps_ratio])
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=42)

    # Overfit: unlimited depth tree
    model_over = DecisionTreeClassifier(max_depth=None, random_state=42)
    model_over.fit(X_train, y_train)

    # Good fit: depth-limited tree
    model_good = DecisionTreeClassifier(max_depth=4, random_state=42)
    model_good.fit(X_train, y_train)

    return {
        "word_count": word_count, "exclamations": exclamations,
        "caps_ratio": caps_ratio, "y": y,
        "X_train": X_train, "y_train": y_train,
        "X_test":  X_test,  "y_test":  y_test,
        "acc_over_train": accuracy_score(y_train, model_over.predict(X_train)),
        "acc_over_test":  accuracy_score(y_test,  model_over.predict(X_test)),
        "acc_good_train": accuracy_score(y_train, model_good.predict(X_train)),
        "acc_good_test":  accuracy_score(y_test,  model_good.predict(X_test)),
        "depth_over": "Unlimited (memorizes everything)",
        "depth_good": "4 (generalizes well)",
    }

# ─── BUILD EXCEL ──────────────────────────────────────────────────────────────

def build_excel():
    lr   = build_linear_regression()
    logr = build_logistic_regression()
    uf   = build_underfitting()
    of   = build_overfitting()

    wb = Workbook()
    wb.remove(wb.active)

    # ── SHEET 1: Overview ─────────────────────────────────────────────────────
    ws = wb.create_sheet("📋 Overview")
    ws.sheet_view.showGridLines = False
    set_col_widths(ws, [3, 28, 22, 22, 22, 22, 3])
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 40

    title = ws.cell(row=2, column=2, value="ML Programs — Visual Reference Guide")
    title.font = Font(name="Arial", bold=True, size=18, color="1A1A2E")
    title.alignment = Alignment(vertical='center')
    ws.merge_cells('B2:F2')
    ws.cell(row=2, column=2).fill = hex_fill("F0F4FF")

    row = 4
    headers = ["Sheet", "Algorithm", "Task", "Dataset", "Key Concept"]
    for c, h in enumerate(headers, 2):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill = hex_fill("1A1A2E")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border()
    ws.row_dimensions[row].height = 20

    rows_data = [
        ["📈 Linear Regression", "Linear Regression", "Regression", "House Size → Price (20 samples)", "Fit a straight line; minimize MSE"],
        ["🎯 Logistic Regression", "Logistic Regression", "Binary Classification", "Study/Sleep hours → Pass/Fail (80 samples)", "Sigmoid → probability → threshold at 0.5"],
        ["📉 Underfitting", "Poly Regression (deg 1 vs 3)", "Regression", "Cubic curve data (60 samples)", "High bias; model too simple for data"],
        ["🔥 Overfitting", "Decision Tree (deep vs shallow)", "Classification", "Spam detection (120 samples)", "Low bias, high variance; memorises noise"],
    ]
    for i, rd in enumerate(rows_data):
        r = row + 1 + i
        bg = "F4F6FB" if i % 2 == 0 else "FFFFFF"
        for c, val in enumerate(rd, 2):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = body_font(size=9)
            cell.fill = hex_fill(bg)
            cell.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
            cell.border = thin_border()
        ws.row_dimensions[r].height = 28

    # concept legend
    row = 10
    ws.cell(row=row, column=2, value="Key Concepts").font = Font(name="Arial", bold=True, size=10, color="1A1A2E")
    ws.row_dimensions[row].height = 18
    concepts = [
        ("Underfitting", "Model too simple → high train AND test error → increase complexity", "FFF3CD", "856404"),
        ("Overfitting",  "Model too complex → low train error, high test error → regularise", "F8D7DA", "721C24"),
        ("Good Fit",     "Low train AND test error, small gap between them → ideal model",    "D4EDDA", "155724"),
        ("R² Score",     "1.0 = perfect fit, 0 = predicts mean, <0 = worse than mean",       "D1ECF1", "0C5460"),
        ("RMSE",         "Root Mean Squared Error — same units as target variable",            "E2D9F3", "4A235A"),
    ]
    for i, (term, desc, bg, fg) in enumerate(concepts):
        r = row + 1 + i
        t = ws.cell(row=r, column=2, value=term)
        t.font = Font(name="Arial", bold=True, size=9, color=fg)
        t.fill = hex_fill(bg)
        t.alignment = Alignment(horizontal='left', indent=1, vertical='center')
        t.border = thin_border()
        d = ws.cell(row=r, column=3, value=desc)
        d.font = Font(name="Arial", size=9, color="1A1A2E")
        d.fill = hex_fill("FAFAFA")
        d.alignment = Alignment(horizontal='left', indent=1, vertical='center', wrap_text=True)
        d.border = thin_border()
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
        ws.row_dimensions[r].height = 22

    # ── SHEET 2: Linear Regression ────────────────────────────────────────────
    ws2 = wb.create_sheet("📈 Linear Regression")
    ws2.sheet_view.showGridLines = False
    set_col_widths(ws2, [2, 16, 16, 16, 16, 2, 20, 20, 2, 20, 20, 2])
    ws2.row_dimensions[1].height = 8

    write_section_title(ws2, 2, 2, "LINEAR REGRESSION  —  House Size → Price Prediction", "1A1A2E", 10)
    ws2.merge_cells('B2:K2')
    ws2.row_dimensions[2].height = 26

    # Explanation
    for r, txt in enumerate([
        "Algorithm: Fits a straight line  y = β₀ + β₁·x  by minimising Mean Squared Error (MSE).",
        "Dataset:   20 houses. Feature = size (sqft). Target = price ($000s).",
        "Formula:   Price = Intercept + (Coefficient × Size)",
    ], 3):
        cell = ws2.cell(row=r, column=2, value=txt)
        cell.font = Font(name="Arial", italic=True, size=9, color="444466")
        ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)
        ws2.row_dimensions[r].height = 16

    # Model metrics
    row = 7
    write_section_title(ws2, row, 2, "MODEL METRICS", "2C3E7A")
    ws2.merge_cells(f'B{row}:E{row}')

    metrics = [
        ("Intercept (β₀)", f"{lr['intercept']:.4f}"),
        ("Coefficient (β₁)", f"{lr['coef']:.4f}"),
        ("Equation", f"Price = {lr['intercept']:.2f} + {lr['coef']:.4f} × Size"),
        ("R² Train", f"{lr['r2_train']:.4f}"),
        ("R² Test",  f"{lr['r2_test']:.4f}"),
        ("RMSE Train", f"${lr['rmse_train']:.2f}k"),
        ("RMSE Test",  f"${lr['rmse_test']:.2f}k"),
    ]
    for i, (k, v) in enumerate(metrics):
        write_kv(ws2, row+1+i, 2, k, v)
        ws2.row_dimensions[row+1+i].height = 18

    # Interpretation
    row2 = row + len(metrics) + 2
    write_section_title(ws2, row2, 2, "INTERPRETATION", "2C3E7A")
    ws2.merge_cells(f'B{row2}:E{row2}')
    notes = [
        f"Every extra sqft adds ~${lr['coef']:.2f}k to price.",
        f"R² of {lr['r2_test']:.2f} means model explains {lr['r2_test']*100:.0f}% of price variance.",
        "Train ≈ Test R² → no overfitting, model generalises well.",
    ]
    for i, n in enumerate(notes):
        c = ws2.cell(row=row2+1+i, column=2, value=f"• {n}")
        c.font = Font(name="Arial", size=9, color="1A1A2E")
        c.fill = hex_fill("F0F4FF")
        c.border = thin_border()
        ws2.merge_cells(start_row=row2+1+i, start_column=2, end_row=row2+1+i, end_column=5)
        ws2.row_dimensions[row2+1+i].height = 16

    # Training data table
    row_t = 7
    write_section_title(ws2, row_t, 7, "TRAINING DATA", "16502A")
    ws2.merge_cells(f'G{row_t}:J{row_t}')
    heads = ["Size (sqft)", "Actual Price ($k)", "Predicted ($k)", "Residual ($k)"]
    for c, h in enumerate(heads, 7):
        cell = ws2.cell(row=row_t+1, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        cell.fill = hex_fill("16502A")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border()

    sort_idx = np.argsort(lr['X_train'])
    for i, idx in enumerate(sort_idx):
        r = row_t + 2 + i
        vals = [lr['X_train'][idx], round(lr['y_train'][idx],2),
                round(lr['y_pred_train'][idx],2),
                round(lr['y_train'][idx]-lr['y_pred_train'][idx],2)]
        for c, v in enumerate(vals, 7):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.font = body_font(size=9)
            cell.fill = hex_fill("F4F6FB" if i%2==0 else "FFFFFF")
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border()
        ws2.row_dimensions[r].height = 16

    # Test data table
    row_tt = row_t + len(lr['X_train']) + 3
    write_section_title(ws2, row_tt, 7, "TEST DATA", "6B2C2C")
    ws2.merge_cells(f'G{row_tt}:J{row_tt}')
    for c, h in enumerate(heads, 7):
        cell = ws2.cell(row=row_tt+1, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        cell.fill = hex_fill("6B2C2C")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border()

    sort_idx2 = np.argsort(lr['X_test'])
    for i, idx in enumerate(sort_idx2):
        r = row_tt + 2 + i
        vals = [lr['X_test'][idx], round(lr['y_test'][idx],2),
                round(lr['y_pred_test'][idx],2),
                round(lr['y_test'][idx]-lr['y_pred_test'][idx],2)]
        for c, v in enumerate(vals, 7):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.font = body_font(size=9)
            cell.fill = hex_fill("FFF0F0" if i%2==0 else "FFFFFF")
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border()
        ws2.row_dimensions[r].height = 16

    # ── SHEET 3: Logistic Regression ──────────────────────────────────────────
    ws3 = wb.create_sheet("🎯 Logistic Regression")
    ws3.sheet_view.showGridLines = False
    set_col_widths(ws3, [2,18,18,18,18,2,16,16,16,16,16,2])
    ws3.row_dimensions[1].height = 8

    write_section_title(ws3, 2, 2, "LOGISTIC REGRESSION  —  Exam Pass/Fail Prediction", "1A1A2E", 10)
    ws3.merge_cells('B2:K2')
    ws3.row_dimensions[2].height = 26

    for r, txt in enumerate([
        "Algorithm: Models P(pass) using σ(z) = 1/(1+e^-z) where z = β₀ + β₁·study + β₂·sleep.",
        "Dataset:   80 students. Features = study hours + sleep hours. Target = pass (1) / fail (0).",
        "Decision:  Predict Pass if P(pass) ≥ 0.5, else Fail.",
    ], 3):
        cell = ws3.cell(row=r, column=2, value=txt)
        cell.font = Font(name="Arial", italic=True, size=9, color="444466")
        ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)
        ws3.row_dimensions[r].height = 16

    row = 7
    write_section_title(ws3, row, 2, "MODEL METRICS", "2C3E7A")
    ws3.merge_cells(f'B{row}:E{row}')

    metrics3 = [
        ("Intercept (β₀)", f"{logr['intercept']:.4f}"),
        ("Coeff — Study hours", f"{logr['coef_study']:.4f}"),
        ("Coeff — Sleep hours", f"{logr['coef_sleep']:.4f}"),
        ("Train Accuracy", f"{logr['acc_train']*100:.1f}%"),
        ("Test Accuracy",  f"{logr['acc_test']*100:.1f}%"),
    ]
    for i, (k, v) in enumerate(metrics3):
        write_kv(ws3, row+1+i, 2, k, v)
        ws3.row_dimensions[row+1+i].height = 18

    # Confusion matrix
    row_cm = row + len(metrics3) + 2
    write_section_title(ws3, row_cm, 2, "CONFUSION MATRIX (Test Set)", "2C3E7A")
    ws3.merge_cells(f'B{row_cm}:E{row_cm}')

    cm = logr['cm']
    cm_labels = [["", "Predicted: FAIL", "Predicted: PASS"],
                 ["Actual: FAIL", cm[0][0], cm[0][1]],
                 ["Actual: PASS", cm[1][0], cm[1][1]]]
    for i, crow in enumerate(cm_labels):
        for j, val in enumerate(crow):
            c = ws3.cell(row=row_cm+1+i, column=2+j, value=val)
            if i == 0 or j == 0:
                c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
                c.fill = hex_fill("2C3E7A")
            elif i == j:  # diagonal (correct)
                c.font = Font(name="Arial", bold=True, size=10, color="155724")
                c.fill = hex_fill("D4EDDA")
            else:          # off-diagonal (errors)
                c.font = Font(name="Arial", bold=True, size=10, color="721C24")
                c.fill = hex_fill("F8D7DA")
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = thin_border()
        ws3.row_dimensions[row_cm+1+i].height = 22

    # Sample data table
    write_section_title(ws3, 7, 7, "SAMPLE TEST DATA (first 20)", "5A2D82")
    ws3.merge_cells('G7:K7')
    t_heads = ["Study Hrs", "Sleep Hrs", "Actual", "Predicted", "Prob(Pass)"]
    for c, h in enumerate(t_heads, 7):
        cell = ws3.cell(row=8, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        cell.fill = hex_fill("5A2D82")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border()

    for i in range(min(20, len(logr['X_test']))):
        r = 9 + i
        vals = [
            round(logr['X_test'][i][0], 2),
            round(logr['X_test'][i][1], 2),
            "Pass" if logr['y_test'][i]==1 else "Fail",
            "Pass" if logr['y_pred_test'][i]==1 else "Fail",
            round(logr['y_prob_test'][i], 3),
        ]
        correct = logr['y_test'][i] == logr['y_pred_test'][i]
        for c, v in enumerate(vals, 7):
            cell = ws3.cell(row=r, column=c, value=v)
            if c == 11:  # last col — colour by prob
                prob = float(v)
                bg = "D4EDDA" if prob > 0.6 else ("F8D7DA" if prob < 0.4 else "FFF3CD")
                cell.fill = hex_fill(bg)
            else:
                cell.fill = hex_fill("F4F6FB" if i%2==0 else "FFFFFF")
            cell.font = body_font(size=9)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border()
        # mark correct/wrong
        cell = ws3.cell(row=r, column=10)
        pill_cell(ws3, r, 10, "✓ Correct" if correct else "✗ Wrong", correct)
        ws3.row_dimensions[r].height = 16

    # ── SHEET 4: Underfitting ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("📉 Underfitting")
    ws4.sheet_view.showGridLines = False
    set_col_widths(ws4, [2,18,14,14,14,14,2,16,16,16,16,2])
    ws4.row_dimensions[1].height = 8

    write_section_title(ws4, 2, 2, "UNDERFITTING  —  Linear Model on Cubic Data", "1A1A2E")
    ws4.merge_cells('B2:K2')
    ws4.row_dimensions[2].height = 26

    for r, txt in enumerate([
        "True pattern: y = x³ − 2x² + x + noise  (cubic/non-linear).",
        "Underfit model: Degree-1 polynomial (straight line) → cannot capture the curve.",
        "Good-fit model: Degree-3 polynomial → matches the true cubic pattern.",
    ], 3):
        cell = ws4.cell(row=r, column=2, value=txt)
        cell.font = Font(name="Arial", italic=True, size=9, color="444466")
        ws4.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)
        ws4.row_dimensions[r].height = 16

    row = 7
    # Comparison table
    write_section_title(ws4, row, 2, "MODEL COMPARISON", "2C3E7A")
    ws4.merge_cells(f'B{row}:F{row}')
    comp_heads = ["Metric", "Underfit (degree 1)", "Good Fit (degree 3)", "What it means"]
    for c, h in enumerate(comp_heads, 2):
        cell = ws4.cell(row=row+1, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        cell.fill = hex_fill("2C3E7A")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border()
        ws4.row_dimensions[row+1].height = 18

    comp_data = [
        ("R² Train",    f"{uf['r2_under_train']:.3f}", f"{uf['r2_good_train']:.3f}", "Higher is better (max 1.0)"),
        ("R² Test",     f"{uf['r2_under_test']:.3f}",  f"{uf['r2_good_test']:.3f}",  "Higher is better (max 1.0)"),
        ("RMSE Train",  f"{uf['rmse_under_train']:.3f}", f"{uf['rmse_good_train']:.3f}", "Lower is better"),
        ("RMSE Test",   f"{uf['rmse_under_test']:.3f}",  f"{uf['rmse_good_test']:.3f}",  "Lower is better"),
        ("Verdict",     "UNDERFIT ✗", "GOOD FIT ✓", "Train R² ≈ Test R² in both cases but values low vs high"),
    ]
    for i, (metric, under, good, meaning) in enumerate(comp_data):
        r = row + 2 + i
        vals = [metric, under, good, meaning]
        bgs  = ["E8EAF6", "FFF3CD", "D4EDDA", "FAFAFA"]
        fgs  = ["1A1A2E", "856404", "155724", "1A1A2E"]
        bolds = [True, True, True, False]
        for c, (v, bg, fg, bold) in enumerate(zip(vals, bgs, fgs, bolds), 2):
            cell = ws4.cell(row=r, column=c, value=v)
            cell.font = Font(name="Arial", bold=bold, size=9, color=fg)
            cell.fill = hex_fill(bg)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border()
        ws4.row_dimensions[r].height = 20

    # Data table
    row_dt = row + len(comp_data) + 3
    write_section_title(ws4, row_dt, 2, "SAMPLE DATA (first 25 training points)", "16502A")
    ws4.merge_cells(f'B{row_dt}:F{row_dt}')
    d_heads = ["x value", "y actual", "Underfit pred", "Good-fit pred", "Underfit error"]
    for c, h in enumerate(d_heads, 2):
        cell = ws4.cell(row=row_dt+1, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        cell.fill = hex_fill("16502A")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border()

    sort_i = np.argsort(uf['X_train'])
    for i, idx in enumerate(sort_i[:25]):
        r = row_dt + 2 + i
        x_v = round(uf['X_train'][idx], 3)
        y_v = round(uf['y_train'][idx], 3)
        u_v = round(uf['y_pred_under_train'][idx], 3)
        g_v = round(uf['y_pred_good_train'][idx], 3)
        err = round(abs(y_v - u_v), 3)
        vals = [x_v, y_v, u_v, g_v, err]
        for c, v in enumerate(vals, 2):
            cell = ws4.cell(row=r, column=c, value=v)
            cell.font = body_font(size=9)
            cell.fill = hex_fill("F4F6FB" if i%2==0 else "FFFFFF")
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border()
        ws4.row_dimensions[r].height = 16

    # Underfitting explanation box
    row_e = 7
    write_section_title(ws4, row_e, 8, "WHY UNDERFITTING HAPPENS", "B7410E")
    ws4.merge_cells(f'H{row_e}:K{row_e}')
    causes = [
        "Model too simple (linear) for complex (cubic) data",
        "Both train AND test errors are HIGH",
        "Adding more data won't help — model can't learn the shape",
        "Fix: increase polynomial degree (model complexity)",
        "Fix: switch to a more expressive model family",
        "Symptom: R² is low on BOTH train and test",
    ]
    for i, c_txt in enumerate(causes):
        r = row_e + 1 + i
        cell = ws4.cell(row=r, column=8, value=f"{'⚠' if i<3 else '✓'} {c_txt}")
        cell.font = Font(name="Arial", size=9,
                         color="856404" if i < 3 else "155724")
        cell.fill = hex_fill("FFF3CD" if i < 3 else "D4EDDA")
        cell.alignment = Alignment(horizontal='left', indent=1, wrap_text=True)
        cell.border = thin_border()
        ws4.merge_cells(start_row=r, start_column=8, end_row=r, end_column=11)
        ws4.row_dimensions[r].height = 18

    # ── SHEET 5: Overfitting ──────────────────────────────────────────────────
    ws5 = wb.create_sheet("🔥 Overfitting")
    ws5.sheet_view.showGridLines = False
    set_col_widths(ws5, [2,20,16,16,16,2,18,16,16,16,16,2])
    ws5.row_dimensions[1].height = 8

    write_section_title(ws5, 2, 2, "OVERFITTING  —  Decision Tree Spam Classifier", "1A1A2E")
    ws5.merge_cells('B2:K2')
    ws5.row_dimensions[2].height = 26

    for r, txt in enumerate([
        "Task: Classify emails as Spam (1) or Not Spam (0).",
        "Features: word count, exclamation marks, caps ratio. Dataset: 120 emails.",
        "Overfit: Decision tree with unlimited depth memorises every training example.",
    ], 3):
        cell = ws5.cell(row=r, column=2, value=txt)
        cell.font = Font(name="Arial", italic=True, size=9, color="444466")
        ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)
        ws5.row_dimensions[r].height = 16

    row = 7
    write_section_title(ws5, row, 2, "MODEL COMPARISON — Accuracy Scores", "2C3E7A")
    ws5.merge_cells(f'B{row}:E{row}')

    acc_heads = ["Metric", "Overfit Tree (∞ depth)", "Good Tree (depth 4)"]
    for c, h in enumerate(acc_heads, 2):
        cell = ws5.cell(row=row+1, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        cell.fill = hex_fill("2C3E7A")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border()

    acc_data = [
        ("Train Accuracy",  f"{of['acc_over_train']*100:.1f}%", f"{of['acc_good_train']*100:.1f}%"),
        ("Test Accuracy",   f"{of['acc_over_test']*100:.1f}%",  f"{of['acc_good_test']*100:.1f}%"),
        ("Train−Test Gap",
         f"{(of['acc_over_train']-of['acc_over_test'])*100:.1f}%",
         f"{(of['acc_good_train']-of['acc_good_test'])*100:.1f}%"),
        ("Tree Depth",      of['depth_over'], of['depth_good']),
        ("Verdict",         "OVERFIT ✗ — memorised training data", "GOOD FIT ✓ — generalises"),
    ]
    for i, row_data in enumerate(acc_data):
        r = row + 2 + i
        metric, over_val, good_val = row_data
        for c, (val, bg, fg) in enumerate(zip(
                [metric, over_val, good_val],
                ["E8EAF6", "F8D7DA", "D4EDDA"],
                ["1A1A2E", "721C24", "155724"]
        ), 2):
            cell = ws5.cell(row=r, column=c, value=val)
            cell.font = Font(name="Arial", bold=(c==2 or i==4), size=9, color=fg)
            cell.fill = hex_fill(bg)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border()
        ws5.row_dimensions[r].height = 20

    # The "tell" explanation
    row_tell = row + len(acc_data) + 3
    write_section_title(ws5, row_tell, 2, "HOW TO SPOT OVERFITTING", "721C24")
    ws5.merge_cells(f'B{row_tell}:E{row_tell}')
    tells = [
        ("Train accuracy = 100% (or near)", "Model has memorised every training example", "F8D7DA", "721C24"),
        ("Large Train−Test gap", "e.g. 100% train vs 72% test → model knows train data only", "F8D7DA", "721C24"),
        ("Fix: Limit tree depth (max_depth=4)", "Restricts memorisation, forces generalisation", "D4EDDA", "155724"),
        ("Fix: Add more training data", "Harder to memorise 10,000 examples than 84", "D4EDDA", "155724"),
        ("Fix: Use Random Forest", "Averaging many trees cancels individual overfit", "D4EDDA", "155724"),
        ("Fix: Apply regularisation (L1/L2)", "Penalise model complexity during training", "D4EDDA", "155724"),
    ]
    for i, (symptom, desc, bg, fg) in enumerate(tells):
        r = row_tell + 1 + i
        s = ws5.cell(row=r, column=2, value=symptom)
        s.font = Font(name="Arial", bold=True, size=9, color=fg)
        s.fill = hex_fill(bg)
        s.alignment = Alignment(horizontal='left', indent=1)
        s.border = thin_border()
        d = ws5.cell(row=r, column=3, value=desc)
        d.font = Font(name="Arial", size=9, color="1A1A2E")
        d.fill = hex_fill("FAFAFA")
        d.alignment = Alignment(horizontal='left', indent=1, wrap_text=True)
        d.border = thin_border()
        ws5.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        ws5.row_dimensions[r].height = 20

    # Sample data table
    write_section_title(ws5, 7, 7, "SAMPLE DATA (first 25 training)", "B7410E")
    ws5.merge_cells('G7:K7')
    d5_heads = ["Word Count", "Exclamations", "Caps Ratio", "Actual", "Label"]
    for c, h in enumerate(d5_heads, 7):
        cell = ws5.cell(row=8, column=c, value=h)
        cell.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        cell.fill = hex_fill("B7410E")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border()

    X_tr = of['X_train']
    y_tr = of['y_train']
    for i in range(min(25, len(X_tr))):
        r = 9 + i
        vals = [int(X_tr[i][0]), int(X_tr[i][1]), round(float(X_tr[i][2]),3),
                int(y_tr[i]), "Spam" if y_tr[i]==1 else "Not Spam"]
        for c, v in enumerate(vals, 7):
            cell = ws5.cell(row=r, column=c, value=v)
            cell.font = body_font(size=9)
            if c == 11:
                cell.fill = hex_fill("F8D7DA" if y_tr[i]==1 else "D4EDDA")
                cell.font = Font(name="Arial", bold=True, size=9,
                                 color="721C24" if y_tr[i]==1 else "155724")
            else:
                cell.fill = hex_fill("FFF5F5" if i%2==0 else "FFFFFF")
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border()
        ws5.row_dimensions[r].height = 16

    # ── SHEET 6: Python Code Reference ────────────────────────────────────────
    ws6 = wb.create_sheet("🐍 Python Code")
    ws6.sheet_view.showGridLines = False
    set_col_widths(ws6, [2, 90, 2])
    ws6.row_dimensions[1].height = 8

    write_section_title(ws6, 2, 2, "PYTHON CODE — All Four Programs", "1A1A2E")
    ws6.row_dimensions[2].height = 26

    code_blocks = [
        ("LINEAR REGRESSION", "16502A", """
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import r2_score, mean_squared_error
import numpy as np

# Sample data: house size (sqft) → price ($k)
X = np.array([600,750,800,850,900,950,1000,1100,1200,1300,
              1400,1500,1600,1700,1800,1900,2000,2200,2400,2600]).reshape(-1,1)
y = 50 + 0.12 * X.flatten() + np.random.normal(0, 15, len(X))

X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=42)

model = LinearRegression()
model.fit(X_train, y_train)

print(f"Equation: Price = {model.intercept_:.2f} + {model.coef_[0]:.4f} x Size")
print(f"R² Train: {r2_score(y_train, model.predict(X_train)):.4f}")
print(f"R² Test:  {r2_score(y_test,  model.predict(X_test)):.4f}")
"""),
        ("LOGISTIC REGRESSION", "2C3E7A", """
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, confusion_matrix
import numpy as np

# Sample data: study hours + sleep hours → pass/fail
study = np.random.uniform(1, 10, 80)
sleep = np.random.uniform(4, 9,  80)
log_odds = -4 + 0.7*study + 0.3*sleep
y = (1/(1 + np.exp(-log_odds)) > 0.5).astype(int)

X = np.column_stack([study, sleep])
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.25, random_state=42)

model = LogisticRegression(max_iter=1000)
model.fit(X_train, y_train)

print(f"Train Accuracy: {accuracy_score(y_train, model.predict(X_train))*100:.1f}%")
print(f"Test  Accuracy: {accuracy_score(y_test,  model.predict(X_test))*100:.1f}%")
print(f"Confusion Matrix:\\n{confusion_matrix(y_test, model.predict(X_test))}")
"""),
        ("UNDERFITTING — Degree 1 vs Degree 3", "B7410E", """
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures
from sklearn.model_selection import train_test_split
from sklearn.metrics import r2_score
import numpy as np

# True pattern is cubic — degree-1 model will underfit
X = np.linspace(-3, 3, 60).reshape(-1,1)
y = X.flatten()**3 - 2*X.flatten()**2 + X.flatten() + np.random.normal(0,2,60)

X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=42)

# Underfitting model (degree 1 = straight line)
lin = LinearRegression().fit(X_train, y_train)
print("=== UNDERFIT (degree 1) ===")
print(f"R² Train: {r2_score(y_train, lin.predict(X_train)):.3f}")   # LOW
print(f"R² Test:  {r2_score(y_test,  lin.predict(X_test)):.3f}")    # LOW

# Good fit model (degree 3 matches true pattern)
poly = PolynomialFeatures(degree=3)
X_tr3 = poly.fit_transform(X_train)
X_te3 = poly.transform(X_test)
good = LinearRegression().fit(X_tr3, y_train)
print("\\n=== GOOD FIT (degree 3) ===")
print(f"R² Train: {r2_score(y_train, good.predict(X_tr3)):.3f}")    # HIGH
print(f"R² Test:  {r2_score(y_test,  good.predict(X_te3)):.3f}")    # HIGH
"""),
        ("OVERFITTING — Unlimited vs Depth-4 Decision Tree", "5A2D82", """
from sklearn.tree import DecisionTreeClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
import numpy as np

# Sample data: email features → spam/not spam
word_count   = np.random.randint(10, 500, 120)
exclamations = np.random.randint(0, 20, 120)
caps_ratio   = np.random.uniform(0, 1, 120)
log_odds = -2 + 0.004*word_count + 0.2*exclamations + 1.5*caps_ratio
y = (1/(1+np.exp(-log_odds)) > 0.5).astype(int)

X = np.column_stack([word_count, exclamations, caps_ratio])
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=42)

# Overfit: unlimited depth memorises every training point
over = DecisionTreeClassifier(max_depth=None, random_state=42).fit(X_train, y_train)
print("=== OVERFIT (unlimited depth) ===")
print(f"Train: {accuracy_score(y_train, over.predict(X_train))*100:.1f}%")  # ~100%
print(f"Test:  {accuracy_score(y_test,  over.predict(X_test))*100:.1f}%")   # much lower

# Good fit: depth 4 generalises
good = DecisionTreeClassifier(max_depth=4, random_state=42).fit(X_train, y_train)
print("\\n=== GOOD FIT (depth 4) ===")
print(f"Train: {accuracy_score(y_train, good.predict(X_train))*100:.1f}%")  # similar
print(f"Test:  {accuracy_score(y_test,  good.predict(X_test))*100:.1f}%")   # similar → generalises
"""),
    ]

    r = 3
    for title, color, code in code_blocks:
        write_section_title(ws6, r, 2, f"# {title}", color)
        ws6.row_dimensions[r].height = 22
        r += 1
        for line in code.split('\n'):
            cell = ws6.cell(row=r, column=2, value=line)
            cell.font = Font(name="Courier New", size=9, color="1A1A2E")
            cell.fill = hex_fill("F8F9FC")
            cell.alignment = Alignment(horizontal='left', indent=1)
            ws6.row_dimensions[r].height = 15
            r += 1
        r += 1

    import os
    path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ML_Programs_Guide.xlsx")
    wb.save(path)
    print(f"Saved: {path}")

build_excel()